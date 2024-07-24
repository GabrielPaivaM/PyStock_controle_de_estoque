import os
import sys
import time
import sqlite3
import datetime
import re

import xlsxwriter
import pandas as pd
import datetime

import openpyxl.drawing.image

from tkinter.filedialog import askdirectory
from tkinter import Tk

from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QMessageBox

from PyQt5 import QtCore, QtGui, QtWidgets

from View.PY.FrmAdmin import Ui_FrmAdmin
from View.PY.FrmLogin import Ui_login
from openpyxl import *

# Definindo os estilos dos botões
button_style = """
    QPushButton {
        background-color: #35cc33;
        color: black;
    }
    QPushButton[selected="true"] {
        background-color: #1a6919;
        color: white;
    }
"""

# Verifique se o arquivo do banco de dados já existe
if os.path.exists('database.db'):
    try:
        # Tente conectar ao banco de dados
        banco = sqlite3.connect('database.db')
        cursor = banco.cursor()
        print("Conexão bem-sucedida ao banco de dados.")

    except sqlite3.DatabaseError as e:
        print("Erro ao conectar ao banco de dados:", e)
        print("Criando um novo banco de dados...")
        os.remove('database.db')  # Remove o arquivo existente
        banco = sqlite3.connect('database.db')
        cursor = banco.cursor()
else:
    print("Criando um novo banco de dados...")
    banco = sqlite3.connect('database.db')
    cursor = banco.cursor()

# Crie tabelas para o banco de dados SQLite
cursor.execute("CREATE TABLE IF NOT EXISTS `clientes` (`CPF` TEXT, `Nome` TEXT, `Endereço` TEXT, `Contato` TEXT, `saldo_devedor` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `fornecedores` (`Nome` TEXT, `Endereço` TEXT, `Contato` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `login` (`usuario` TEXT, `senha` TEXT, `nivel` TEXT, `nome` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `monitoramento_vendas` (`vendedor` TEXT, `cliente` TEXT, `qtde_vendido` TEXT, `total_venda` TEXT, `horario_venda` TEXT, `cpf_da_venda` TEXT,`a_prazo` INTEGER, `produto` TEXT, 'total_custo' TEXT, 'forma_pagamento' TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `monitoramento_compras` (`comprador` TEXT, `fornecedor` TEXT, `qtde_comprado` TEXT, `total_compra` TEXT, `horario_compra` TEXT, `produto` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `produtos` (`cód_produto` TEXT, `descrição` TEXT, `valor_unitário` TEXT, `qtde_estoque` TEXT, `fornecedor` TEXT, `valor_de_custo` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `quem_vendeu_mais` (`nome` TEXT, `total_qtde` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `vendas` (`cód` TEXT, `produto` TEXT, `valor_unitário` TEXT, `qtde` TEXT, `total` TEXT, `id` INTEGER, `valor_de_custo` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `compras` (`cód` TEXT, `produto` TEXT, `valor_de_custo` TEXT, `qtde` TEXT, `total` TEXT, `id` INTEGER, `fornecedor` TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `sangrias` (`valor` TEXT, `motivo` TEXT, 'data' TEXT);")
cursor.execute("CREATE TABLE IF NOT EXISTS `valor_inicio_dia` (`valor` TEXT);")



cursor.execute("SELECT COUNT(*) FROM 'login'")
result = cursor.fetchone()[0]
if result == 0:
    cursor.execute("INSERT INTO `login` (`usuario`, `senha`, `nivel`, `nome`) VALUES ('1', '1', 'admin', 'Renato');")

# Comite as transações para garantir que elas sejam concluídas
banco.commit()

class FrmLogin(QMainWindow):

    def __init__(self):

        QMainWindow.__init__(self)

        self.ui = Ui_login()
        self.ui.setupUi(self)

        self.setWindowTitle("PyStock")
        self.setWindowIcon(QIcon('LogoIco.ico'))


        # Botão de logar no sistema
        self.ui.pushButton.clicked.connect(lambda: self.logar())

    def logar(self):

        global window, UserLogado, senhaLogada

        # Pegando os colaboradores cadastrados no banco
        cursor.execute("SELECT * FROM login")
        logins = cursor.fetchall()

        # Pegando as informações inseridas
        usuario = self.ui.lineEdit.text()
        senha = self.ui.lineEdit_2.text()

        # Verificando cada Usuário
        for login in logins:

            if usuario != login[0]:
                self.ui.lineEdit.setStyleSheet('background-color: rgba(0, 0 , 0, 0);border: 2px solid rgba(0,0,0,0);'
                                               'border-bottom-color: rgb(255, 17, 49);color: rgb(0,0,0);padding-bottom: 8px;'
                                               'border-radius: 0px;font: 10pt "Montserrat";')

            if senha != login[1]:
                self.ui.lineEdit_2.setStyleSheet('background-color: rgba(0, 0 , 0, 0);border: 2px solid rgba(0,0,0,0);'
                                                 'border-bottom-color: rgb(255, 17, 49);color: rgb(0,0,0);padding-bottom: 8px;'
                                                 'border-radius: 0px;font: 10pt "Montserrat";')

            # Caso os dados estejam no banco, é iniciado o Frm de acordo com o nivel
            if usuario == login[0] and senha == login[1]:

                UserLogado = login[3]
                senhaLogada = login[1]

                self.ui.lineEdit.setStyleSheet('background-color: rgba(0, 0 , 0, 0);border: 2px solid rgba(0,0,0,0);'
                                               'border-bottom-color: rgb(54, 204, 51);color: rgb(0,0,0);padding-bottom: 8px;'
                                               'border-radius: 0px;font: 10pt "Montserrat";')

                self.ui.lineEdit_2.setStyleSheet('background-color: rgba(0, 0 , 0, 0);border: 2px solid rgba(0,0,0,0);'
                                                 'border-bottom-color: rgb(54, 204, 51);color: rgb(0,0,0);padding-bottom: 8px;'
                                                 'border-radius: 0px;font: 10pt "Montserrat";')

                window.close()
                window = FrmAdmin()
                window.setWindowFlags(window.windowFlags() | QtCore.Qt.Window)
                window.showMaximized()
                window.show()
                break


class FrmAdmin(QMainWindow):

    def __init__(self):
        global filtro, search_fornecedores, window

        QMainWindow.__init__(self)

        self.ui = Ui_FrmAdmin()
        self.ui.setupUi(self)

        self.setWindowTitle("PyStock")
        self.setWindowIcon(QIcon('LogoIco.ico'))

        self.ui.line_codigo_produto_cadastrar.hide()
        # self.ui.line_codigo_alterar_produto.hide()

        # Definindo botões do menu
        self.menu_buttons = {
            'home': self.ui.btn_home,
            'colaboradores': self.ui.btn_colaboradores,
            'vendas_monitoramento': self.ui.btn_vendas_monitoramento,
            'compras_monitoramento': self.ui.btn_compras_monitoramento,
            'clientes': self.ui.btn_clientes,
            'vendas': self.ui.btn_vendas,
            'compras': self.ui.btn_compras,
            'fornecedores': self.ui.btn_fornecedores,
            'produtos': self.ui.btn_produtos,
            'configs': self.ui.btn_configs
        }

        for button in self.menu_buttons.values():
            button.setStyleSheet(button_style)

        # Nome do Usuário
        self.ui.lbl_seja_bem_vindo.setText(f'Seja Bem-Vindo(a) - {UserLogado}')
        self.ui.lbl_titulo_vendas.setText(f'Vendedor(a) - {UserLogado}')
        self.ui.lbl_seja_bem_vindo.setFixedWidth(500)

        # Configurando páginas e os botões do menu

        # Home
        self.ui.btn_home.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_home))

        # Colaboradores
        self.ui.btn_colaboradores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_colaboradores))
        self.ui.btn_cadastrar_colaboradores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_cadastro_colaboradores))
        self.ui.btn_alterar_colaboradores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_alterar_colaboradores))

        self.ui.btn_sangria.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_gerar_sangria))

        self.ui.btn_cadastro_colaboradores.clicked.connect(self.CadastroColaboradores)
        self.ui.btn_finalizar_alterar_colaboradores.clicked.connect(self.AlterarColaboradores)

        self.ui.btn_valor_ok.clicked.connect(self.SalvarValorDeInicio)
        self.ui.btn_finalizar_dia.clicked.connect(self.GerarXlsFinalDia)

        self.ui.line_senha_alterar_colaboradores.setEchoMode(QLineEdit.EchoMode.Password)
        self.ui.btn_exluir_colaboradores.clicked.connect(self.ExcluirColaboradores)
        self.ui.tabela_alterar_colaboradores.doubleClicked.connect(self.setTextAlterarColaboradores)

        # Botões para ver/esconder senha inserida
        self.ui.btn_ver_senha_cadastro_colaboradores.clicked.connect(self.VerSenhaCadastroColaboradores)
        self.ui.btn_ver_senha_alterar_colaboradores.clicked.connect(self.VerSenhaAlterarColaboradores)

        self.ui.btn_ver_senha_gerar_sangria.clicked.connect(self.VerSenhaCadastroSangria)

        # Tabela pg_colaboradores
        self.ui.tabela_colaboradores.setColumnWidth(0, 467)
        self.ui.tabela_colaboradores.setColumnWidth(1, 467)
        self.ui.tabela_colaboradores.setColumnWidth(2, 467)

        # Tabela alterar_colaboradores
        self.ui.tabela_alterar_colaboradores.setColumnWidth(0, 584)
        self.ui.tabela_alterar_colaboradores.setColumnWidth(1, 584)
        self.ui.tabela_alterar_colaboradores.setColumnWidth(2, 584)

        # Monitoramento vendas
        self.ui.btn_vendas_monitoramento.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_monitoramento_vendas))

        # Tabela Monitoramento vendas
        self.ui.tabela_monitoramento_vendas.setColumnWidth(0, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(1, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(2, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(3, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(4, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(5, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(6, 175)
        self.ui.tabela_monitoramento_vendas.setColumnWidth(7, 176)


        # Monitoramento compras
        self.ui.btn_compras_monitoramento.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_monitoramento_compras))

        # Tabela Monitoramento compras
        self.ui.tabela_monitoramento_compras.setColumnWidth(0, 234)
        self.ui.tabela_monitoramento_compras.setColumnWidth(1, 234)
        self.ui.tabela_monitoramento_compras.setColumnWidth(2, 234)
        self.ui.tabela_monitoramento_compras.setColumnWidth(3, 234)
        self.ui.tabela_monitoramento_compras.setColumnWidth(4, 234)
        self.ui.tabela_monitoramento_compras.setColumnWidth(5, 234)

        # Clientes
        self.ui.btn_clientes.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_clientes))
        self.ui.btn_cadastrar_clientes_clientes.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_cadastrar_clientes))
        self.ui.btn_alterar_clientes_clientes.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_alterar_clientes))
        self.ui.btn_finalizar_cadastro_clientes.clicked.connect(self.CadastrarClientes)
        self.ui.btn_exclui_clientes.clicked.connect(self.ExcluirClientes)
        self.ui.tabela_alterar_clientes.doubleClicked.connect(self.setTextAlterarClientes)
        self.ui.btn_finalizar_alteracao_alterar_clientes.clicked.connect(self.AlterarClientes)

        # Tabela Clientes
        self.ui.tabela_clientes.setColumnWidth(0, 280)
        self.ui.tabela_clientes.setColumnWidth(1, 280)
        self.ui.tabela_clientes.setColumnWidth(2, 280)
        self.ui.tabela_clientes.setColumnWidth(3, 280)
        self.ui.tabela_clientes.setColumnWidth(4, 280)

        # Tabela Cadastrar Clientes
        self.ui.tabela_cadastrar_clientes.setColumnWidth(0, 438)
        self.ui.tabela_cadastrar_clientes.setColumnWidth(1, 438)
        self.ui.tabela_cadastrar_clientes.setColumnWidth(2, 438)
        self.ui.tabela_cadastrar_clientes.setColumnWidth(3, 438)

        # Tabela Alterar Clientes
        self.ui.tabela_alterar_clientes.setColumnWidth(0, 438)
        self.ui.tabela_alterar_clientes.setColumnWidth(1, 438)
        self.ui.tabela_alterar_clientes.setColumnWidth(2, 438)
        self.ui.tabela_alterar_clientes.setColumnWidth(3, 438)

        # Vendas
        self.ui.btn_vendas.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_vendas))

        # Tabela Vendas
        self.ui.tabela_vendas.setColumnWidth(0, 234)
        self.ui.tabela_vendas.setColumnWidth(1, 234)
        self.ui.tabela_vendas.setColumnWidth(2, 234)
        self.ui.tabela_vendas.setColumnWidth(3, 234)
        self.ui.tabela_vendas.setColumnWidth(4, 234)
        self.ui.tabela_vendas.setColumnWidth(5, 234)

        # Compras
        self.ui.btn_compras.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_compras))

        # Tabela compras
        self.ui.tabela_vendas_carregamento.setColumnWidth(0, 234)
        self.ui.tabela_vendas_carregamento.setColumnWidth(1, 234)
        self.ui.tabela_vendas_carregamento.setColumnWidth(2, 234)
        self.ui.tabela_vendas_carregamento.setColumnWidth(3, 234)
        self.ui.tabela_vendas_carregamento.setColumnWidth(4, 234)
        self.ui.tabela_vendas_carregamento.setColumnWidth(5, 234)

        # Fornecedores
        self.ui.btn_fornecedores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_fornecedores))
        self.ui.btn_adicionar_forncedores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_cadastrar_fornecedores))
        self.ui.btn_editar_fornecedores.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_alterar_fornecedores))
        self.ui.btn_cadastrar_forncedores.clicked.connect(self.CadastrarFornecedores)
        self.ui.tabela_alterar_fornecedores.doubleClicked.connect(self.setTextAlterarFornecedores)
        self.ui.btn_alterar_fornecedores.clicked.connect(self.AlterarFornecedores)
        self.ui.btn_excluir_fornecedores.clicked.connect(self.ExluirFornecedores)

        # Tabela Fornecedores
        self.ui.tabela_fornecedores.setColumnWidth(0, 466)
        self.ui.tabela_fornecedores.setColumnWidth(1, 466)
        self.ui.tabela_fornecedores.setColumnWidth(2, 467)

        # Tabela Cadastrar Fornecedores
        self.ui.tabela_cadastrar_fornecedores.setColumnWidth(0, 584)
        self.ui.tabela_cadastrar_fornecedores.setColumnWidth(1, 584)
        self.ui.tabela_cadastrar_fornecedores.setColumnWidth(2, 584)

        # Tabela Alterar Fornecedores
        self.ui.tabela_alterar_fornecedores.setColumnWidth(0, 584)
        self.ui.tabela_alterar_fornecedores.setColumnWidth(1, 584)
        self.ui.tabela_alterar_fornecedores.setColumnWidth(2, 584)

        # Produtos
        self.ui.btn_produtos.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_produtos))
        self.ui.btn_cadastrar_produto.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_cadastar_produtos))
        self.ui.btn_alterar_produto.clicked.connect(
            lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_alterar_produtos))
        self.ui.btn_finalizar_cadastro_produtos.clicked.connect(self.CadastrarProdutos)
        self.ui.btn_excluir_produto.clicked.connect(self.ExcluirProdutos)
        self.ui.tabela_alterar_produto.doubleClicked.connect(self.setTextAlterarProdutos)
        self.ui.btn_finalizar_alterar_produto.clicked.connect(self.AlterarProdutos)

        # Tabela Produtos
        self.ui.tabela_produto.setColumnWidth(0, 200)
        self.ui.tabela_produto.setColumnWidth(1, 200)
        self.ui.tabela_produto.setColumnWidth(2, 200)
        self.ui.tabela_produto.setColumnWidth(3, 200)
        self.ui.tabela_produto.setColumnWidth(4, 200)
        self.ui.tabela_produto.setColumnWidth(5, 200)
        self.ui.tabela_produto.setColumnWidth(6, 200)

        # Tabela Cadastrar Produtos
        self.ui.tabela_cadastro_produto.setColumnWidth(0, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(1, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(2, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(3, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(4, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(5, 250)
        self.ui.tabela_cadastro_produto.setColumnWidth(6, 250)


        # Tabela Alterar Produtos
        self.ui.tabela_alterar_produto.setColumnWidth(0, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(1, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(2, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(3, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(4, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(5, 250)
        self.ui.tabela_alterar_produto.setColumnWidth(6, 250)

        # Configurações
        self.ui.btn_configs.clicked.connect(lambda: self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_configuracoes))

        # Voltar
        self.ui.btn_voltar.clicked.connect(self.Voltar)

        # Atualizando Tabelas
        self.AtualizaTabelasLogin()
        self.AtualizaTabelasClientes()
        self.AtualizaTabelasFornecedores()
        self.AtualizaTabelasProdutos()
        self.AtualizaTabelaVendas()
        self.AtualizaTabelaMonitoramentoVendas()
        self.AtualizaTabelaCompras()
        self.AtualizaTabelaMonitoramentoCompras()

        # iniciando Hora e Data do Sistema
        tempo = QTimer(self)
        tempo.timeout.connect(self.HoraData)
        tempo.timeout.connect(self.Sair)
        tempo.start(1000)

        # Atuliza previsão de texto das barras de pesquisa
        self.AtualizaCompleterSearchFornecedores()
        self.AtualizaCompleterSearchProdutos()
        self.AtualizaCompleterSearchColaboradores()
        self.AtualizaCompleterSearchClientes()
        self.AtualizaCompleterSearchVendas()
        self.AtualizaCompleterSearchCompras()
        self.AtualizarLBLValorInicio()

        # Conectando com a função fazer a pesquisa dos dados inseridos

        # Produtos
        self.ui.btn_pesquisar_produto.clicked.connect(lambda: self.SearchProdutos(pg='Produtos'))
        self.ui.line_search_Bar_produtos.returnPressed.connect(lambda: self.SearchProdutos(pg='Produtos'))

        self.ui.line_search_Bar_alterar_produto.returnPressed.connect(lambda: self.SearchProdutos(pg='Alterar'))
        self.ui.btn_pesquisar_alterar_produto.clicked.connect(lambda: self.SearchProdutos(pg='Alterar'))

        self.ui.line_search_Bar_cadastrar_produto.returnPressed.connect(lambda: self.SearchProdutos(pg='Cadastrar'))
        self.ui.btn_pesquisar_cadastrar_produto.clicked.connect(lambda: self.SearchProdutos(pg='Cadastrar'))

        # Fornecedores
        self.ui.btn_filtrar_fornecedores.clicked.connect(lambda: self.SearchFornecedores(pg='Fornecedores'))
        self.ui.line_search_Bar_fornecedores.returnPressed.connect(lambda: self.SearchFornecedores(pg='Fornecedores'))

        self.ui.btn_flitrar_alterar_fornecedores.clicked.connect(lambda: self.SearchFornecedores(pg='Alterar'))
        self.ui.line_search_Bar_altarar_fornecedor.returnPressed.connect(lambda: self.SearchFornecedores(pg='Alterar'))

        self.ui.btn_pesquisar_cadastrar_fornecedores.clicked.connect(lambda: self.SearchFornecedores(pg='Cadastrar'))
        self.ui.line_search_Bar_cadastrar_fornecedores.returnPressed.connect(
            lambda: self.SearchFornecedores(pg='Cadastrar'))

        # Colaboradores
        self.ui.btn_pesquisar_colaboradores.clicked.connect(lambda: self.SearchColaboradores(pg='Colaboradores'))
        self.ui.line_search_bar_colaboradores.returnPressed.connect(
            lambda: self.SearchColaboradores(pg='Colaboradores'))

        self.ui.btn_pesquisar_alterar_colaboradores.clicked.connect(lambda: self.SearchColaboradores(pg='Alterar'))
        self.ui.line_search_bar_buscar_alterar_colaboradores.returnPressed.connect(
            lambda: self.SearchColaboradores(pg='Alterar'))

        # Clientes
        self.ui.btn_pesquisar_clientes.clicked.connect(lambda: self.SearchClientes(pg='Clientes'))
        self.ui.line_search_Bar_clientes.returnPressed.connect(lambda: self.SearchClientes(pg='Clientes'))

        self.ui.btn_filtrar_alterar_clientes.clicked.connect(lambda: self.SearchClientes(pg='Alterar'))
        self.ui.line_search_Bar_alterar_clientes.returnPressed.connect(lambda: self.SearchClientes(pg='Alterar'))

        self.ui.btn_pesquisar_cadastro_clientes.clicked.connect(lambda: self.SearchClientes(pg='Cadastrar'))
        self.ui.line_search_Bar_cadastrar_clientes.returnPressed.connect(lambda: self.SearchClientes(pg='Cadastrar'))

        self.ui.line_cpf_cadastrar_clientes.setMaxLength(14)
        self.ui.line_contato_cadastrar_clientes.setMaxLength(16)

        self.ui.line_alterar_cpf_alterar_clientes.setMaxLength(14)
        self.ui.line_alterar_contato_alterar_clientes.setMaxLength(16)

        self.ui.line_cadastrar_contato_fornecedores.setMaxLength(16)
        self.ui.line_alterar_contato_fornecedor.setMaxLength(16)

        # Formatando número de contato dos Fornecedores
        self.ui.line_cadastrar_contato_fornecedores.textChanged.connect(
            lambda: self.FormataNumeroContato(pg='CadastrarFornecedores'))
        self.ui.line_alterar_contato_fornecedor.textChanged.connect(
            lambda: self.FormataNumeroContato(pg='AlterarFornecedores'))

        # Formatando número de contato dos Clientes
        self.ui.line_contato_cadastrar_clientes.textChanged.connect(lambda: self.FormataNumeroContato(pg='CadastrarClientes'))
        self.ui.line_alterar_contato_alterar_clientes.textChanged.connect(lambda: self.FormataNumeroContato(pg='AlterarClientes'))

        # Formatando CPF dos Clientes
        self.ui.line_cpf_cadastrar_clientes.textChanged.connect(lambda: self.FormataCPFClientes(pg='Cadastrar'))
        self.ui.line_alterar_cpf_alterar_clientes.textChanged.connect(lambda: self.FormataCPFClientes(pg='Alterar'))
        self.ui.line_cliente.textChanged.connect(lambda: self.FormataCPFClientes(pg='Vendas'))
        self.ui.line_xls_cpf_clientes.textChanged.connect(lambda: self.FormataCPFClientes(pg='clientes'))
        self.ui.line_xls_cpf_clientes.setMaxLength(14)

        # Formatando o valor de produtos
        self.ui.line_valor_cadastrar_produto_produto.textChanged.connect(lambda: self.FormataValorProduto(pg='Cadastrar'))
        self.ui.line_valorcusto_cadastrar_produto.textChanged.connect(lambda: self.FormataValorProduto(pg= 'CadastrarValorCusto'))
        self.ui.line_valor_alterar_produto.textChanged.connect(lambda: self.FormataValorProduto(pg='Alterar'))
        self.ui.line_valorcusto_alterar_produto.textChanged.connect(lambda: self.FormataValorProduto(pg= 'AlterarValorCusto'))

        self.ui.line_valor_inicio_dia.textChanged.connect(lambda: self.FormataValorProduto(pg= 'ValorInicio'))
        self.ui.line_valor_gerar_sangria.textChanged.connect(lambda: self.FormataValorProduto(pg= 'ValorSangria'))


        # Formatando valor desconto
        self.ui.line_desconto_vendas.setMaxLength(3)
        self.ui.line_desconto_vendas.setText("0")

        # Pesquisando produto pelo código
        self.ui.line_codigo_produto.returnPressed.connect(self.PesquisandoProdutoPeloCodigo)
        self.ui.btn_confirmar_codigo.clicked.connect(self.PesquisandoProdutoPeloCodigo)
        self.ui.line_search_bar_vendas.returnPressed.connect(self.CodProdutoVendas)
        self.ui.line_search_bar_vendas_carregamento.returnPressed.connect(self.CodProdutoCompras)


        # Vendas
        self.ui.btn_adicionar_compra.clicked.connect(self.CadastrandoVendas)
        self.ui.btn_excluir_item.clicked.connect(self.ExcluirVenda)

        # Compras
        self.ui.btn_adicionar_carregamento.clicked.connect(self.CadastrandoCompras) #implementar
        self.ui.btn_excluir_item_carregamento.clicked.connect(self.ExcluirCompra)


        # Conectando com a função Troco
        self.ui.line_troco.textChanged.connect(self.Troco)
        self.ui.line_troco.returnPressed.connect(self.Troco)

        self.ui.btn_cadastro_sangria.clicked.connect(self.SalvarSangria)

        # Conectando com a função de formatar a data
        self.ui.line_data_monitoramento_vendas.setMaxLength(7)
        self.ui.line_data_monitoramento_vendas.textChanged.connect(self.formatar_data)

        self.ui.line_data_monitoramento_compras.setMaxLength(7)
        self.ui.line_data_monitoramento_compras.textChanged.connect(self.formatar_data_compras)

        # Atualizando Total
        self.AtualizaTotal()
        self.AtualizaTotalCompras()

        # Conectando com função de Finalizar a venda
        self.ui.btn_finalizar_compra.clicked.connect(self.FinalizarVendas)
        self.AtualizaTabelaMonitoramentoVendas()

        self.ui.btn_finalizar_carregamento.clicked.connect(self.FinalizarCompras)
        self.AtualizaTabelaMonitoramentoCompras()

        # Conectando com a função de Limpar a tabela do monitamento
        self.ui.btn_limpar_tabela_monitoramento_vendas.clicked.connect(self.LimparTabelaMonitoramento)
        self.ui.btn_limpar_tabela_compras.clicked.connect(self.LimparTabelaMonitoramentoCompras)

        # Conectando com a barra de pesquisa do monitoramento
        self.ui.line_search_bar_monitoramentoto_vendas_vendas_vendas.returnPressed.connect(self.SearchMonitoramentoVendas)
        self.ui.btn_filtrar_monitoramento_vendas.clicked.connect(self.SearchMonitoramentoVendas)

        self.ui.line_search_bar_monitoramentoto_vendas_vendas_compras.returnPressed.connect(self.SearchMonitoramentoCompras)
        self.ui.btn_filtrar_monitoramento_compras.clicked.connect(self.SearchMonitoramentoCompras)

        # Conectando com a função de gerar o arquivo xlsx de vendas
        self.ui.btn_gerar_excel_monitoramento_vendas.clicked.connect(self.GerarXls)

        # Conectando com a função de gerar o arquivo xlsx de pendencias
        self.ui.btn_gerarxls_pendencia_clientes.clicked.connect(self.GerarXlsPendencias)

        # Conectando com a função de limpar pendencias
        self.ui.btn_limpar_pendencia_cliente.clicked.connect(self.LimparPendencias)

        # Conectando com a função de gerar o arquivo xlsx de compras
        self.ui.btn_gerar_excel_compras.clicked.connect(self.GerarXlsCompras)

        # Função para encerrar o programa após 20m
        self.ui.btn_salvar.clicked.connect(self.Futuro)
        self.ui.btn_salvar.clicked.connect(self.Sair)

        # Configurar a aba inicial para "Home"
        self.ui.Telas_do_menu.setCurrentWidget(self.ui.pg_home)
        self.change_page('home', self.ui.pg_home)

    def change_page(self, button_key, page):
        self.ui.Telas_do_menu.setCurrentWidget(page)
        self.update_button_styles(button_key)

    def update_button_styles(self, active_button_key):
        for key, button in self.menu_buttons.items():
            button.setProperty("selected", key == active_button_key)
            button.style().unpolish(button)
            button.style().polish(button)
            button.update()

    # Pequenas Funções
    def Voltar(self):
        global window

        window.close()
        window = FrmLogin()
        window.show()

    def HoraData(self):
        tempoAtual = QTime.currentTime()
        tempoTexto = tempoAtual.toString('hh:mm:ss')
        data_atual = datetime.date.today()
        dataTexto = data_atual.strftime('%d/%m/%Y')

        # Colaboradores
        self.ui.lbl_hora_data_colaboradores.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_alterar_colaboradores.setText(f'{dataTexto} {tempoTexto}')

        # Monitoramento de Vendas e compras
        self.ui.lbl_hora_data_monitoramento_vendas.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_monitoramento_compras.setText(f'{dataTexto} {tempoTexto}')

        # Vendas e compras
        self.ui.lbl_hora_data.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_carregamento.setText(f'{dataTexto} {tempoTexto}')

        # Produtos
        self.ui.lbl_hora_data_produtos.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_alterar_produto.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_cadastrar_produto.setText(f'{dataTexto} {tempoTexto}')

        # Fornecedores
        self.ui.lbl_hora_data_fornecedores.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_alterar_fornecedores.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_cadastrar_fornecedores.setText(f'{dataTexto} {tempoTexto}')

        # Clientes
        self.ui.lbl_hora_data_clientes.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_cadastrar_clientes.setText(f'{dataTexto} {tempoTexto}')
        self.ui.lbl_hora_data_alterar_clientes.setText(f'{dataTexto} {tempoTexto}')

    def PesquisandoProdutoPeloCodigo(self):
        produtos = list()

        cod_inserido = self.ui.line_codigo_produto

        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        produtos.clear()
        tabela = self.ui.tabela_produto

        for produto in banco_produtos:
            produtos.append(produto[0])

        items = tabela.findItems(cod_inserido.text(), Qt.MatchExactly)

        if items:
            item = items[0]
            tabela.setCurrentItem(item)

            cod_inserido.setStyleSheet(StyleNormal)

        else:
            cod_inserido.setStyleSheet(StyleError)

    def CodProdutoVendas(self):
        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        produto_inserido = self.ui.line_search_bar_vendas

        for pos, produto in enumerate(banco_produtos):
            if produto[1] == produto_inserido.text():
                self.ui.line_codigo_vendas.setText(produto[0])
                break

    def CodProdutoCompras(self):
        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        produto_inserido = self.ui.line_search_bar_vendas_carregamento

        for pos, produto in enumerate(banco_produtos):
            if produto[1] == produto_inserido.text():
                self.ui.line_codigo_vendas_carregamento.setText(produto[0])
                break

    def ConfirmarCliente(self):
        global search_clientes

        cliente = self.ui.line_cliente

        if cliente.text() in search_clientes:
            cliente.setStyleSheet('''
                background-color: rgba(0, 0 , 0, 0);
                border: 2px solid rgba(0,0,0,0);
                border-bottom-color: rgb(15, 168, 103);
                color: rgb(0,0,0);
                padding-bottom: 8px;
                border-radius: 0px;
                font: 10pt "Montserrat";''')
        else:
            cliente.setStyleSheet(StyleError)

    def FinalizarVendas(self):
        print("FinalizarVendas iniciado")  # Adicionado para depuração
        a_prazo_check = self.ui.radio_venda_prazo
        pix_check = self.ui.radio_venda_pix
        dinheiro_check = self.ui.radio_venda_dinheiro
        debito_check = self.ui.radio_venda_debito
        credito_check = self.ui.radio_venda_credito
        cliente = self.ui.line_cliente.text()

        forma_de_pagamento = ""

        if pix_check.isChecked():
            forma_de_pagamento = "PIX"

        if dinheiro_check.isChecked():
            forma_de_pagamento = "Dinheiro"

        if debito_check.isChecked():
            forma_de_pagamento = "Débito"

        if credito_check.isChecked():
            forma_de_pagamento = "Crédito"

        if a_prazo_check.isChecked():
            forma_de_pagamento = "Não informado"

        if not a_prazo_check.isChecked() and not credito_check.isChecked() and not debito_check.isChecked() and not dinheiro_check.isChecked() and not pix_check.isChecked():
            self.Popup('Vendas', 'Nenhuma forma de pagamento informada')
            return

        cursor.execute('SELECT * FROM vendas')
        banco_vendas = cursor.fetchall()
        print(f"Banco de vendas: {banco_vendas}")  # Adicionado para depuração

        cursor.execute('SELECT * FROM quem_vendeu_mais')
        banco_quem_mais_vendeu = cursor.fetchall()

        cursor.execute('SELECT cpf FROM clientes WHERE nome = ?', (cliente,))
        cpf_da_venda = cursor.fetchone()

        tempoAtual = QTime.currentTime()
        tempoTexto = tempoAtual.toString('hh:mm:ss')
        data_atual = datetime.date.today()
        dataTexto = data_atual.strftime('%d/%m/%Y')

        vendedor = UserLogado
        data_hora = f'{dataTexto} / {tempoTexto}'

        if cliente != '':
            if banco_vendas and cpf_da_venda is not None and a_prazo_check.isChecked() and cliente != '':
                a_prazo = 1

                for venda in banco_vendas:
                    qtde_vendido = venda[3]
                    totalVenda = venda[4]
                    totalCusto = venda[6]
                    produto = venda[1]

                    comando_SQL = 'INSERT INTO monitoramento_vendas VALUES (?,?,?,?,?,?,?,?,?,?)'
                    dados = (vendedor, cliente, qtde_vendido, totalVenda, data_hora, cpf_da_venda[0], a_prazo, produto, totalCusto, forma_de_pagamento)
                    print(f"Dados para monitoramento: {dados}")  # Adicionado para depuração
                    cursor.execute(comando_SQL, dados)
                    banco.commit()

                cursor.execute('SELECT saldo_devedor FROM clientes WHERE nome = ?', (cliente,))
                saldo_devedor = cursor.fetchone()

                saldo_devedor_total = float(''.join(re.findall(r'\d', saldo_devedor[0]))) / 100 + sum(
                    [float(venda[4]) for venda in banco_vendas])
                saldo_devedor_convertido = 'R$' + "{:.2f}".format(saldo_devedor_total)

                cursor.execute('UPDATE clientes SET saldo_devedor = ? WHERE nome = ?', (saldo_devedor_convertido, cliente,))
                banco.commit()

                colaboradores = [colaborador[0] for colaborador in banco_quem_mais_vendeu]

                if vendedor in colaboradores:
                    colaborador = next(colab for colab in banco_quem_mais_vendeu if colab[0] == vendedor)
                    cursor.execute(
                        'UPDATE quem_vendeu_mais SET total_qtde = ? WHERE nome = ?',
                        (int(colaborador[1]) + sum([int(venda[3]) for venda in banco_vendas]), vendedor)
                    )
                    banco.commit()
                else:
                    comando_SQL = 'INSERT INTO quem_vendeu_mais VALUES (?,?)'
                    dados = (vendedor, sum([int(venda[3]) for venda in banco_vendas]))
                    cursor.execute(comando_SQL, dados)
                    banco.commit()

                for venda in banco_vendas:
                    cursor.execute("UPDATE produtos SET qtde_estoque = qtde_estoque - ? WHERE cód_produto = ?",
                                   (venda[3], venda[0]))
                    banco.commit()
            elif banco_vendas and not a_prazo_check.isChecked() and cliente != '':
                a_prazo = 0

                for venda in banco_vendas:
                    qtde_vendido = venda[3]
                    totalVenda = venda[4]
                    totalCusto = venda[6]
                    produto = venda[1]

                    comando_SQL = 'INSERT INTO monitoramento_vendas VALUES (?,?,?,?,?,?,?,?,?,?)'
                    dados = (vendedor, cliente, qtde_vendido, totalVenda, data_hora, None, a_prazo, produto, totalCusto, forma_de_pagamento)
                    print(f"Dados para monitoramento: {dados}")  # Adicionado para depuração
                    cursor.execute(comando_SQL, dados)
                    banco.commit()

                colaboradores = [colaborador[0] for colaborador in banco_quem_mais_vendeu]

                if vendedor in colaboradores:
                    colaborador = next(colab for colab in banco_quem_mais_vendeu if colab[0] == vendedor)
                    cursor.execute(
                        'UPDATE quem_vendeu_mais SET total_qtde = ? WHERE nome = ?',
                        (int(colaborador[1]) + sum([int(venda[3]) for venda in banco_vendas]), vendedor)
                    )
                    banco.commit()
                else:
                    comando_SQL = 'INSERT INTO quem_vendeu_mais VALUES (?,?)'
                    dados = (vendedor, sum([int(venda[3]) for venda in banco_vendas]))
                    cursor.execute(comando_SQL, dados)
                    banco.commit()

                for venda in banco_vendas:
                    cursor.execute("UPDATE produtos SET qtde_estoque = qtde_estoque - ? WHERE cód_produto = ?",
                                   (venda[3], venda[0]))
                    banco.commit()

            elif cpf_da_venda is None and a_prazo_check.isChecked() and cliente != '':
                self.Popup('Vendas', 'Não foi possível achar o CPF do cliente informado')
            elif not banco_vendas:
                self.Popup('Vendas', 'Nenhum produto informado')

            cursor.execute('DELETE FROM vendas')
            banco.commit()
            self.AtualizaTabelaVendas()
            self.AtualizaTotal()
            self.AtualizaTabelasClientes()
            self.AtualizaTabelasProdutos()
            self.AtualizaTabelaMonitoramentoVendas()
            self.AtualizaCompleterSearchVendas()
            self.AtualizaCompleterSearchCompras()
        else:
            self.Popup('Vendas', 'Nenhum cliente informado')

        self.ui.line_codigo_vendas.clear()
        self.ui.line_cliente.clear()
        self.ui.line_quantidade_vendas.clear()
        self.ui.line_desconto_vendas.setText("0")
        self.ui.lbl_troco.setText('Troco: R$ 0,00')
        self.ui.line_cliente.setStyleSheet(StyleNormal)
        self.ui.line_search_bar_vendas.clear()
        self.ui.line_troco.clear()
        self.ui.radio_venda_prazo.setChecked(False)
        self.ui.radio_venda_pix.setChecked(False)
        self.ui.radio_venda_dinheiro.setChecked(False)
        self.ui.radio_venda_debito.setChecked(False)
        self.ui.radio_venda_credito.setChecked(False)
        self.ui.line_desconto_vendas.setStyleSheet(StyleNormal)
        self.ui.line_codigo_vendas.setStyleSheet(StyleNormal)
        self.ui.line_quantidade_vendas.setStyleSheet(StyleNormal)

    def FinalizarCompras(self):
        cursor.execute('SELECT * FROM compras')
        banco_compras = cursor.fetchall()

        tempoAtual = QTime.currentTime()
        tempoTexto = tempoAtual.toString('hh:mm:ss')
        data_atual = datetime.date.today()
        dataTexto = data_atual.strftime('%d/%m/%Y')

        comprador = UserLogado
        data_hora = f'{dataTexto} / {tempoTexto}'

        if banco_compras:
            for compra in banco_compras:
                fornecedor = compra[6]
                qtde_comprado = compra[3]
                totalCompra = compra[4]
                produto = compra[1]

                comando_SQL = 'INSERT INTO monitoramento_compras VALUES (?,?,?,?,?,?)'
                dados = f'{comprador}', f'{fornecedor}', f'{qtde_comprado}', f'{totalCompra}', f'{data_hora}', f'{produto}'
                cursor.execute(comando_SQL, dados)
                banco.commit()

            for compra in banco_compras:
                cursor.execute("UPDATE produtos SET qtde_estoque = qtde_estoque + ? WHERE cód_produto = ?",
                               (compra[3], compra[0]))
                banco.commit()
        elif not banco_compras:
            self.Popup('Compras', 'Nenhum produto informado')

        cursor.execute('DELETE FROM compras')
        banco.commit()
        self.AtualizaTabelaCompras()
        self.AtualizaTotalCompras()
        self.AtualizaTabelasClientes()
        self.AtualizaTabelasProdutos()
        self.AtualizaTabelaMonitoramentoCompras()
        self.AtualizaCompleterSearchCompras()

        self.ui.line_codigo_vendas_carregamento.clear()
        self.ui.line_quantidade_vendas_carregamento.clear()
        self.ui.line_search_bar_vendas_carregamento.clear()
        self.ui.line_codigo_vendas_carregamento.setStyleSheet(StyleNormal)
        self.ui.line_quantidade_vendas_carregamento.setStyleSheet(StyleNormal)

    def Troco(self):
        cursor.execute('SELECT * FROM vendas')
        banco_vendas = cursor.fetchall()

        troco_desejado = self.ui.line_troco.text()
        troco_desejado = ''.join(filter(str.isdigit, troco_desejado))  # Apenas dígitos

        if troco_desejado.isnumeric():
            troco_desejado_int = int(troco_desejado)

            # Converte os valores de venda para float e soma
            vendas = [float(venda[4].replace(',', '.')) for venda in banco_vendas]
            total_vendas = sum(vendas)

            # Calcula o troco
            troco = troco_desejado_int / 100 - total_vendas

            # Formata o troco para exibição
            troco_formatado = f"{troco:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.ui.lbl_troco.setText("Troco: R$ " + troco_formatado)

            # Atualiza line_troco com o valor formatado
            troco_desejado_formatado = f"{troco_desejado_int / 100:,.2f}".replace(",", "X").replace(".", ",").replace(
                "X", ".")
            self.ui.line_troco.setText( troco_desejado_formatado)
            self.ui.line_troco.setCursorPosition(len(troco_desejado_formatado))
        else:
            # Limpa a linha de texto se a entrada não for um número válido
            self.ui.line_troco.setText('')

    def AtualizaTotal(self):
        # Obtém todas as vendas do banco de dados
        cursor.execute('SELECT * FROM vendas')
        banco_vendas = cursor.fetchall()

        vendas = list()
        vendascusto = list()

        for venda in banco_vendas:
            vendat = ''.join(re.findall(r'\d', venda[4]))
            vendas.append(int(vendat))

            vendatcusto = ''.join(re.findall(r'\d', venda[6]))  # Assumindo que o custo está na coluna 5
            vendascusto.append(int(vendatcusto))

        # Calcula o total das vendas e dos custos
        total_vendas = sum(vendas) * 0.01
        total_custos = sum(vendascusto) * 0.01

        # Calcula o lucro e a porcentagem de lucro em relação ao total de vendas
        lucro = total_vendas - total_custos
        porcentagem_lucro = (lucro / total_vendas) * 100 if total_vendas != 0 else 0

        # Atualiza a interface do usuário com os novos valores
        self.ui.lbl_total_valor.setText(f'Total: R$ {total_vendas:.2f}')
        self.ui.lbl_lucro.setText(f'Lucro: R$ {lucro:.2f} ({porcentagem_lucro:.2f}%)')

        self.Troco()  # Chama o método Troco se necessário

    def AtualizaTotalCompras(self):

        cursor.execute('SELECT * FROM compras')
        banco_compras = cursor.fetchall()

        compras = list()

        for pos, compra in enumerate(banco_compras):
            comprat = ''.join(re.findall(r'\d', compra[4]))
            compras.append(int(comprat))

        total = lang.toString(sum(compras) * 0.01, 'f', 2)
        self.ui.lbl_total_valor_carregamento.setText(f'{total}')
        self.Troco()

    def Futuro(self):
        global futuroTexto
        atual = datetime.datetime.now()

        futuro = atual + datetime.timedelta(minutes=20)
        futuroTexto = futuro.time().strftime('%H:%M:%S')

    def Sair(self):
        global futuroTexto
        tempoAtual = QTime.currentTime()
        tempoTexto = tempoAtual.toString('hh:mm:ss')

        if self.ui.checkBox_finalizar_app.isChecked() == True:
            if tempoTexto == futuroTexto:
                sys.exit()

    def AtualizarLBLValorInicio(self):
        cursor.execute('SELECT valor FROM valor_inicio_dia')
        valor = cursor.fetchall()

        if valor:

            valor_formatado = float(valor[0][0]) * 0.01

            self.ui.lbl_valor_inicio.setText(f'Valor de inicio: R$ {valor_formatado:.2f}')
            return

        self.ui.lbl_valor_inicio.setText(f'Valor de inicio: R$ 0,00')

    def SalvarValorDeInicio(self):
        valor = ''.join(re.findall(r'\d', self.ui.line_valor_inicio_dia.text()))

        if valor == '':
            self.Popup("Configuraçoes", "Adicione um valor")
            return

        cursor.execute('DELETE FROM valor_inicio_dia')
        banco.commit()

        comando_SQL = 'INSERT INTO valor_inicio_dia VALUES(?)'
        dados = (valor,)
        cursor.execute(comando_SQL, dados)
        banco.commit()

        self.ui.line_valor_inicio_dia.setText("")
        self.AtualizarLBLValorInicio()

    def SalvarSangria(self):
        valor = ''.join(re.findall(r'\d', self.ui.line_valor_gerar_sangria.text()))
        motivo = self.ui.line_motivo_gerar_sangria.text()
        senhaDada = self.ui.line_senha_gerar_sangria.text()
        senhaCorreta = senhaLogada

        tempoAtual = QTime.currentTime()
        data_atual = datetime.date.today()
        dataTexto = data_atual.strftime('%d/%m/%Y')

        data_hora = f'{dataTexto}'

        if valor == '' or motivo == '' or senhaDada == '':
            self.Popup("Gerar Sangria", "Preencha todos os campos")
            return

        if not senhaDada == senhaCorreta:
            self.Popup("Gerar Sangria", "Senha de usuario incorreta")
            return

        comando_SQL = 'INSERT INTO sangrias VALUES(?,?,?)'
        dados = f'{valor}', f'{motivo}', f'{data_hora}'
        cursor.execute(comando_SQL, dados)
        banco.commit()

        print(f'salvo valor {valor} pelo motivo: {motivo} na data {data_hora}')

        self.ui.line_senha_gerar_sangria.setText("")
        self.ui.line_motivo_gerar_sangria.setText("")
        self.ui.line_valor_gerar_sangria.setText("")
        self.Popup("Gerar Sangria", "Sangria cadastrada com sucesso")

    def LimparTabelaMonitoramento(self):
        id = self.ui.tabela_monitoramento_vendas.currentRow()

        if id != -1:
            cursor.execute('SELECT rowid, * FROM monitoramento_vendas ORDER BY rowid ASC')
            banco_monitoramento = cursor.fetchall()

            if id < len(banco_monitoramento):
                venda_selecionada = banco_monitoramento[id]
                id_deletado = venda_selecionada[0]  # Usando rowid para identificar a venda a ser deletada
                cliente = venda_selecionada[2]  # Nome do cliente na terceira posição do resultado da consulta
                total_venda = float(venda_selecionada[4])  # Valor da venda na quinta posição do resultado da consulta
                a_prazo = venda_selecionada[7]  # Status "A Prazo" na oitava posição do resultado da consulta

                if a_prazo == 1:  # Verificando se a venda foi "A Prazo"
                    cursor.execute('SELECT saldo_devedor FROM clientes WHERE nome = ?', (cliente,))
                    saldo_devedor = cursor.fetchone()

                    if saldo_devedor:
                        saldo_devedor_atual = float(''.join(re.findall(r'\d', saldo_devedor[0]))) / 100
                        novo_saldo_devedor = saldo_devedor_atual - total_venda

                        saldo_devedor_convertido = 'R$' + "{:.2f}".format(novo_saldo_devedor)
                        cursor.execute('UPDATE clientes SET saldo_devedor = ? WHERE nome = ?',
                                       (saldo_devedor_convertido, cliente))
                        banco.commit()

                cursor.execute('DELETE FROM monitoramento_vendas WHERE rowid = ?', (id_deletado,))
                banco.commit()

            self.AtualizaTabelaMonitoramentoVendas()
            self.AtualizaCompleterSearchVendas()
            self.AtualizaTabelasClientes()

    def LimparTabelaMonitoramentoCompras(self):
        id = self.ui.tabela_monitoramento_compras.currentRow()

        if id != -1:
            cursor.execute('SELECT rowid, * FROM monitoramento_compras ORDER BY rowid ASC')
            banco_monitoramento = cursor.fetchall()

            if id < len(banco_monitoramento):
                compra_selecionada = banco_monitoramento[id]
                id_deletado = compra_selecionada[0]  # Usando rowid para identificar a venda a ser deletada
                produto = compra_selecionada[6]  # Nome do cliente na terceira posição do resultado da consulta
                quantidade = int(compra_selecionada[3])  # Valor da venda na quinta posição do resultado da consulta

                cursor.execute('SELECT qtde_estoque FROM produtos WHERE descrição = ?', (produto,))
                qtde_estoque = cursor.fetchone()

                if qtde_estoque:
                    quantidade_atual = qtde_estoque[0]
                    nova_quantidade = int(quantidade_atual) - quantidade

                    cursor.execute('UPDATE produtos SET qtde_estoque = ? WHERE descrição = ?',
                                       (nova_quantidade, produto))
                    banco.commit()

                cursor.execute('DELETE FROM monitoramento_compras WHERE rowid = ?', (id_deletado,))
                banco.commit()

            self.AtualizaTabelaMonitoramentoCompras()
            self.AtualizaCompleterSearchCompras()
            self.AtualizaTabelasProdutos()


    # Popups
    def Popup(self, titulo, texto):
        msg = QMessageBox()
        msg.setWindowTitle(titulo)
        msg.setText(texto)

        icon = QIcon()
        icon.addPixmap(QPixmap("View/Imagens/Logo Ico.ico"), QIcon.Normal, QIcon.Off)
        msg.setWindowIcon(icon)
        x = msg.exec_()

    def PopupXlsDiretorio(self):
        msg = QMessageBox()
        msg.setWindowTitle("Erro - Gerar Excel")
        msg.setText('Selecione um diretório válido!')

        icon = QIcon()
        icon.addPixmap(QPixmap("View/Imagens/Logo Ico.ico"), QIcon.Normal, QIcon.Off)
        msg.setWindowIcon(icon)
        x = msg.exec_()

    def PopupXls(self):
        msg = QMessageBox()
        msg.setWindowTitle("Erro - Gerar Excel")
        msg.setText('Verifique se não há um ARQUIVO com o mesmo nome aberto!')

        icon = QIcon()
        icon.addPixmap(QPixmap("View/Imagens/Logo Ico.ico"), QIcon.Normal, QIcon.Off)
        msg.setWindowIcon(icon)
        x = msg.exec_()

    def PoupXlsBancoVazio(self):
        msg = QMessageBox()
        msg.setWindowTitle("Erro - Gerar Excel")
        msg.setText('Nenhuma venda informada!')

        icon = QIcon()
        icon.addPixmap(QPixmap("View/Imagens/Logo Ico.ico"), QIcon.Normal, QIcon.Off)
        msg.setWindowIcon(icon)
        x = msg.exec_()

    def PoupXlsBancoVazioCompras(self):
        msg = QMessageBox()
        msg.setWindowTitle("Erro - Gerar Excel")
        msg.setText('Nenhuma compra informada!')

        icon = QIcon()
        icon.addPixmap(QPixmap("View/Imagens/Logo Ico.ico"), QIcon.Normal, QIcon.Off)
        msg.setWindowIcon(icon)
        x = msg.exec_()

    # Função para formatar a data desejada
    def formatar_data(self):
        text = self.ui.line_data_monitoramento_vendas.text()

        # Remover quaisquer caracteres que não sejam números
        digits_only = ''.join(filter(str.isdigit, text))

        # Adicionar uma barra após os primeiros dois caracteres (mês)
        formatted_text = digits_only[:2] if len(digits_only) >= 2 else digits_only
        if len(digits_only) > 2:
            formatted_text += '/' + digits_only[2:]

        # Atualizar o campo de texto
        self.ui.line_data_monitoramento_vendas.setText(formatted_text)
        self.ui.line_data_monitoramento_vendas.setCursorPosition(len(formatted_text))

    def formatar_data_compras(self):
        text = self.ui.line_data_monitoramento_compras.text()

        # Remover quaisquer caracteres que não sejam números
        digits_only = ''.join(filter(str.isdigit, text))

        # Adicionar uma barra após os primeiros dois caracteres (mês)
        formatted_text = digits_only[:2] if len(digits_only) >= 2 else digits_only
        if len(digits_only) > 2:
            formatted_text += '/' + digits_only[2:]

        # Atualizar o campo de texto
        self.ui.line_data_monitoramento_compras.setText(formatted_text)
        self.ui.line_data_monitoramento_compras.setCursorPosition(len(formatted_text))

    def GerarXlsPendencias(self):
        cpf = self.ui.line_xls_cpf_clientes.text()

        cursor.execute('SELECT cpf FROM clientes WHERE cpf = ?', (cpf,))
        vefCPF = cursor.fetchall()

        if cpf != '' and vefCPF:
            cursor.execute(
                'SELECT vendedor, cliente, qtde_vendido, total_venda, horario_venda, produto FROM monitoramento_vendas WHERE cpf_da_venda = ? AND a_prazo = 1',
                (cpf,))
            banco_monitoramento = cursor.fetchall()

            if len(banco_monitoramento) > 0:
                Tk().withdraw()
                diretorio = askdirectory()

                if diretorio != '':
                    try:
                        file_path = f'{diretorio}/Pendencias_{cpf}.xlsx'
                        workbook = xlsxwriter.Workbook(file_path)
                        worksheet = workbook.add_worksheet("Pendências")

                        # Define formats
                        header_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'green'})
                        currency_format = workbook.add_format({'num_format': 'R$ #,##0.00'})

                        # Write headers
                        headers = ["Vendedor", "Cliente", "Quantidade Vendida", "Valor Total", "Data da Venda",
                                   "Produto"]
                        for col_num, header in enumerate(headers):
                            worksheet.write(0, col_num, header, header_format)

                        # Write data
                        total_devido = 0
                        for row_num, venda in enumerate(banco_monitoramento, start=1):
                            worksheet.write(row_num, 0, venda[0])
                            worksheet.write(row_num, 1, venda[1])
                            worksheet.write(row_num, 2, venda[2])
                            worksheet.write(row_num, 3, venda[3], currency_format)
                            worksheet.write(row_num, 4, venda[4])
                            worksheet.write(row_num, 5, venda[5])
                            total_devido += float(venda[3])

                        # Write total devido
                        worksheet.write(row_num + 1, 2, "Total Devido", header_format)
                        worksheet.write(row_num + 1, 3, total_devido, currency_format)

                        workbook.close()

                    except Exception as e:
                        print("Error generating XLSX:", e)
                        self.PopupXls()
                else:
                    self.PopupXlsDiretorio()
            else:
                self.Popup("Erro - Gerar Excel", "Nenhuma pendencia para este cliente")
        else:
            self.Popup("Erro - Gerar Excel", "CPF não cadastrado")

    def GerarXls(self):
        dataDesejada = self.ui.line_data_monitoramento_vendas.text()
        dataDesejada_formatted = dataDesejada.replace('/', '-')  # Replace '/' with '-' for valid file name

        cursor.execute(
            'SELECT vendedor, cliente, qtde_vendido, total_venda, horario_venda, produto, total_custo, forma_pagamento FROM monitoramento_vendas WHERE SUBSTR(horario_venda, 4, 7) = ?',
            (dataDesejada,))
        banco_monitoramento = cursor.fetchall()

        if len(banco_monitoramento) > 0:
            Tk().withdraw()
            diretorio = askdirectory()

            if diretorio != '':
                try:
                    file_path = f'{diretorio}/Relatório_Vendas_{dataDesejada_formatted}.xlsx'
                    workbook = xlsxwriter.Workbook(file_path)
                    worksheet = workbook.add_worksheet("Relatório")

                    # Define formats
                    header_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'green'})
                    currency_format = workbook.add_format({'num_format': 'R$ #,##0.00'})
                    percentage_format = workbook.add_format({'num_format': '0.00%'})

                    # Write headers
                    headers = ["Vendedor", "Cliente", "Quantidade Vendida", "Valor Total", "Data da Venda", "Produto",
                               "Custo", "Lucro", "Porcentagem de Lucro", "Forma de Pagamento"]
                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header, header_format)

                    # Write data
                    total_faturado = 0
                    total_custos = 0
                    for row_num, venda in enumerate(banco_monitoramento, start=1):
                        valor_total = float(venda[3])
                        custo_total = float(venda[6]) / 100  # Dividindo por 100 para corrigir o formato
                        lucro = valor_total - custo_total
                        porcentagem_lucro = (lucro / valor_total) if valor_total != 0 else 0

                        worksheet.write(row_num, 0, venda[0])
                        worksheet.write(row_num, 1, venda[1])
                        worksheet.write(row_num, 2, venda[2])
                        worksheet.write(row_num, 3, valor_total, currency_format)
                        worksheet.write(row_num, 4, venda[4])
                        worksheet.write(row_num, 5, venda[5])
                        worksheet.write(row_num, 6, custo_total, currency_format)
                        worksheet.write(row_num, 7, lucro, currency_format)
                        worksheet.write(row_num, 8, porcentagem_lucro, percentage_format)
                        worksheet.write(row_num, 9, venda[7])

                        total_faturado += valor_total
                        total_custos += custo_total

                    # Write total faturado
                    worksheet.write(row_num + 1, 2, "Total Faturado", header_format)
                    worksheet.write(row_num + 1, 3, total_faturado, currency_format)

                    # Calculate and write lucro and porcentagem_lucro
                    lucro_total = total_faturado - total_custos
                    porcentagem_lucro_total = (lucro_total / total_faturado) if total_faturado != 0 else 0
                    worksheet.write(row_num + 2, 2, "Lucro Total", header_format)
                    worksheet.write(row_num + 2, 3, lucro_total, currency_format)
                    worksheet.write(row_num + 3, 2, "Porcentagem de Lucro Total", header_format)
                    worksheet.write(row_num + 3, 3, porcentagem_lucro_total, percentage_format)

                    workbook.close()

                except Exception as e:
                    print("Error generating XLSX:", e)
                    self.PopupXls()
            else:
                self.PopupXlsDiretorio()
        else:
            self.PoupXlsBancoVazio()

    def GerarXlsFinalDia(self):
        try:
            data_atual = datetime.date.today()
            dataTexto = data_atual.strftime('%d-%m-%Y')  # Troquei '/' por '-'

            # Obtém o valor de início do caixa
            cursor.execute('SELECT valor FROM valor_inicio_dia')
            valorDeInicio = cursor.fetchone()
            valorDeInicio = float(valorDeInicio[0]) * 0.01 if valorDeInicio else 0.0

            # Obtém as vendas do dia
            cursor.execute(
                'SELECT total_venda FROM monitoramento_vendas WHERE SUBSTR(horario_venda, 1, 10) = ?',
                (dataTexto.replace('-', '/'),))  # Ajuste para consulta de data no banco de dados
            banco_monitoramento = cursor.fetchall()

            # Obtém as sangrias do dia
            cursor.execute(
                'SELECT valor, motivo FROM sangrias WHERE SUBSTR(data, 1, 10) = ?',
                (dataTexto.replace('-', '/'),))  # Ajuste para consulta de data no banco de dados
            banco_sangrias = cursor.fetchall()

            total_vendas = sum(float(venda[0]) for venda in banco_monitoramento)
            total_sangrias = sum(float(sangria[0]) * 0.01 for sangria in banco_sangrias)

            valor_final_caixa = (valorDeInicio + total_vendas) - total_sangrias

            Tk().withdraw()
            diretorio = askdirectory()

            if diretorio:
                try:
                    file_path = f'{diretorio}/Relatório_Sangrias_{dataTexto}.xlsx'

                    # Check if file already exists and rename it if necessary
                    base, extension = os.path.splitext(file_path)
                    count = 1
                    while os.path.exists(file_path):
                        file_path = f'{base}_{count}{extension}'
                        count += 1

                    workbook = xlsxwriter.Workbook(file_path)
                    worksheet = workbook.add_worksheet("Relatório")

                    # Define formats
                    header_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'green'})
                    currency_format = workbook.add_format({'num_format': 'R$ #,##0.00'})

                    # Write valores iniciais
                    worksheet.write(0, 0, "Valor de Início do Caixa", header_format)
                    worksheet.write(0, 1, valorDeInicio, currency_format)

                    worksheet.write(1, 0, "Total de Vendas do Dia", header_format)
                    worksheet.write(1, 1, total_vendas, currency_format)

                    # Write sangrias data
                    sangrias_headers = ["Valor", "Motivo"]
                    for col_num, header in enumerate(sangrias_headers):
                        worksheet.write(3, col_num, header, header_format)

                    row_num = 4
                    for sangria in banco_sangrias:
                        valor_sangria = float(sangria[0]) * 0.01
                        worksheet.write(row_num, 0, valor_sangria, currency_format)
                        worksheet.write(row_num, 1, sangria[1])
                        row_num += 1

                    # Write total sangrias
                    worksheet.write(row_num, 0, "Total Sangrias", header_format)
                    worksheet.write(row_num, 1, total_sangrias, currency_format)

                    # Write valor final do caixa
                    worksheet.write(row_num + 2, 0, "Valor Final do Caixa", header_format)
                    worksheet.write(row_num + 2, 1, valor_final_caixa, currency_format)

                    workbook.close()

                except Exception as e:
                    print("Error generating XLSX:", e)
                    self.PopupXls()
            else:
                self.PopupXlsDiretorio()

        except Exception as e:
            print("Error:", e)
            self.PopupXls()

    def LimparPendencias(self):
        cpf = self.ui.line_xls_cpf_clientes.text()

        pix_check = self.ui.radio_venda_pix_clientes
        dinheiro_check = self.ui.radio_venda_dinheiro_clientes
        debito_check = self.ui.radio_venda_debito_clientes
        credito_check = self.ui.radio_venda_credito_clientes
        cliente = self.ui.line_cliente.text()

        forma_de_pagamento = ""

        if pix_check.isChecked():
            forma_de_pagamento = "PIX"

        if dinheiro_check.isChecked():
            forma_de_pagamento = "Dinheiro"

        if debito_check.isChecked():
            forma_de_pagamento = "Débito"

        if credito_check.isChecked():
            forma_de_pagamento = "Crédito"

        if not credito_check.isChecked() and not debito_check.isChecked() and not dinheiro_check.isChecked() and not pix_check.isChecked():
            self.Popup('Clientes', 'Nenhuma forma de pagamento informada')
            return

        if cpf != '':
            cursor.execute('UPDATE monitoramento_vendas SET forma_pagamento = ? WHERE a_prazo = 1 AND cpf_da_venda = ?',
                           (forma_de_pagamento, cpf,))
            banco.commit()

            cursor.execute('UPDATE monitoramento_vendas SET a_prazo = 0 WHERE a_prazo = 1 AND cpf_da_venda = ?',(cpf,))
            banco.commit()

            cursor.execute('UPDATE clientes SET saldo_devedor = "R$ 0,00" WHERE CPF = ?',(cpf,))
            banco.commit()

            self.Popup("Limpar pendencias", "Todas pendencias foram quitadas")
        else:
            self.Popup("Limpar pendencias", "Nenhum CPF informado")

        self.ui.radio_venda_pix_clientes.setChecked(False)
        self.ui.radio_venda_dinheiro_clientes.setChecked(False)
        self.ui.radio_venda_debito_clientes.setChecked(False)
        self.ui.radio_venda_credito_clientes.setChecked(False)

        self.AtualizaTabelasClientes()
        self.AtualizaTabelaMonitoramentoVendas()

    def GerarXlsCompras(self):
        dataDesejada = self.ui.line_data_monitoramento_compras.text()
        dataDesejada_formatted = dataDesejada.replace('/', '-')  # Replace '/' with '-' for valid file name

        cursor.execute(
            'SELECT comprador, fornecedor, qtde_comprado, total_compra, horario_compra, produto FROM monitoramento_compras WHERE SUBSTR(horario_compra, 4, 7) = ?',
            (dataDesejada,))
        banco_monitoramento = cursor.fetchall()

        if len(banco_monitoramento) > 0:
            Tk().withdraw()
            diretorio = askdirectory()

            if diretorio != '':
                try:
                    file_path = f'{diretorio}/Relatório_Compras_{dataDesejada_formatted}.xlsx'
                    workbook = xlsxwriter.Workbook(file_path)
                    worksheet = workbook.add_worksheet("Relatório")

                    # Define formats
                    header_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'green'})
                    currency_format = workbook.add_format({'num_format': 'R$ #,##0.00'})

                    # Write headers
                    headers = ["Comprador", "Forncedor", "Quantidade Comprado", "Valor Total", "Data da Compra", "Produto"]
                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header, header_format)

                    # Write data
                    total_comprado = 0
                    for row_num, compra in enumerate(banco_monitoramento, start=1):
                        worksheet.write(row_num, 0, compra[0])
                        worksheet.write(row_num, 1, compra[1])
                        worksheet.write(row_num, 2, compra[2])
                        worksheet.write(row_num, 3, compra[3], currency_format)
                        worksheet.write(row_num, 4, compra[4])
                        worksheet.write(row_num, 5, compra[5])
                        total_comprado += float(compra[3])

                    # Write total faturado
                    worksheet.write(row_num + 1, 2, "Total Comprado", header_format)
                    worksheet.write(row_num + 1, 3, total_comprado, currency_format)

                    workbook.close()

                except Exception as e:
                    print("Error generating XLSX:", e)
                    self.PopupXls()
            else:
                self.PopupXlsDiretorio()
        else:
            self.PoupXlsBancoVazioCompras()

    # Função para formartar o número de contato inserido
    def FormataNumeroContato(self, pg):
        global numero

        if pg == 'CadastrarFornecedores':
            numero = self.ui.line_cadastrar_contato_fornecedores
        if pg == 'AlterarFornecedores':
            numero = self.ui.line_alterar_contato_fornecedor
        if pg == 'CadastrarClientes':
            numero = self.ui.line_contato_cadastrar_clientes
        if pg == 'AlterarClientes':
            numero = self.ui.line_alterar_contato_alterar_clientes

        texto = numero.text()
        tamanho = len(numero.text())

        if tamanho == 1 and texto.isnumeric() == True:
            numero.setText(f'({texto}')

        if tamanho == 3 and texto[1:].isnumeric() == True:
            numero.setText(f'{texto}) ')

        if tamanho == 6 and texto[5].isnumeric() == True:
            numero.setText(f'{texto} ')

        if tamanho == 11 and texto[7].isnumeric() == True:
            numero.setText(f'{texto}-')

    # Função para formatar o cpf inserido
    def FormataCPFClientes(self, pg):

        if pg == 'clientes':
            CPF = self.ui.line_xls_cpf_clientes
        if pg == 'Cadastrar':
            CPF = self.ui.line_cpf_cadastrar_clientes
        if pg == 'Alterar':
            CPF = self.ui.line_alterar_cpf_alterar_clientes
        if pg == 'Vendas':
            CPF = self.ui.line_cliente

        TextoInserido = CPF.text()
        TamanhoDoTexto = len(CPF.text())

        if TamanhoDoTexto == 3 and TextoInserido.isnumeric() == True:
            CPF.setText(f'{TextoInserido}.')
        if TamanhoDoTexto == 7 and TextoInserido[4:].isnumeric() == True:
            CPF.setText(f'{TextoInserido}.')
        if TamanhoDoTexto == 11 and TextoInserido[8:].isnumeric() == True:
            CPF.setText(f'{TextoInserido}-')

    # Função para formatar o valor inserido
    def FormataValorProduto(self, pg):

        if pg == 'Cadastrar':
            Valor = self.ui.line_valor_cadastrar_produto_produto.text()
        if pg == 'Alterar':
            Valor = self.ui.line_valor_alterar_produto.text()
        if pg == 'Desconto':
            Valor = self.ui.line_desconto_vendas.text()
        if pg == 'CadastrarValorCusto':
            Valor = self.ui.line_valorcusto_cadastrar_produto.text()
        if pg == 'AlterarValorCusto':
            Valor = self.ui.line_valorcusto_alterar_produto.text()
        if pg == 'ValorInicio':
            Valor = self.ui.line_valor_inicio_dia.text()
        if pg == 'ValorSangria':
            Valor = self.ui.line_valor_gerar_sangria.text()

        # Remove qualquer caractere que não seja numérico
        valorInserido = ''.join(filter(str.isdigit, Valor))

        if valorInserido:

            # Converte para inteiro e depois para float para tratar corretamente
            valor = int(valorInserido)
            valor_float = valor / 100

            # Formata como valor monetário
            numeroFormatado = f"{valor_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            if pg == 'Cadastrar':
                self.ui.line_valor_cadastrar_produto_produto.setText(numeroFormatado)
                # Move o cursor para o final
                self.ui.line_valor_cadastrar_produto_produto.setCursorPosition(len(numeroFormatado))
            if pg == 'Alterar':
                self.ui.line_valor_alterar_produto.setText(numeroFormatado)
                # Move o cursor para o final
                self.ui.line_valor_alterar_produto.setCursorPosition(len(numeroFormatado))
            if pg == 'Desconto':
                self.ui.line_desconto_vendas.setText(numeroFormatado)
                # Move o cursor para o final
                self.ui.line_desconto_vendas.setCursorPosition(len(numeroFormatado))
            if pg == 'CadastrarValorCusto':
                self.ui.line_valorcusto_cadastrar_produto.setText(numeroFormatado)

                self.ui.line_valorcusto_cadastrar_produto.setCursorPosition(len(numeroFormatado))

            if pg == 'AlterarValorCusto':
                self.ui.line_valorcusto_alterar_produto.setText(numeroFormatado)

                self.ui.line_valorcusto_alterar_produto.setCursorPosition(len(numeroFormatado))

            if pg == 'ValorInicio':
                self.ui.line_valor_inicio_dia.setText(numeroFormatado)

                self.ui.line_valor_inicio_dia.setCursorPosition(len(numeroFormatado))

            if pg == 'ValorSangria':
                self.ui.line_valor_gerar_sangria.setText(numeroFormatado)

                self.ui.line_valor_gerar_sangria.setCursorPosition(len(numeroFormatado))

    # Função para Fazer a pesquisa dos item inserido
    def SearchProdutos(self, pg):
        tabela = self
        produto = self

        if pg == 'Produtos':
            tabela = self.ui.tabela_produto
            produto = self.ui.line_search_Bar_produtos

        if pg == 'Alterar':
            tabela = self.ui.tabela_alterar_produto
            produto = self.ui.line_search_Bar_alterar_produto

        if pg == 'Cadastrar':
            tabela = self.ui.tabela_cadastro_produto
            produto = self.ui.line_search_Bar_cadastrar_produto

        items = tabela.findItems(produto.text(), Qt.MatchContains)

        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    def SearchFornecedores(self, pg):
        tabela = self
        produto = self

        if pg == 'Fornecedores':
            tabela = self.ui.tabela_fornecedores
            produto = self.ui.line_search_Bar_fornecedores

        if pg == 'Alterar':
            tabela = self.ui.tabela_alterar_fornecedores
            produto = self.ui.line_search_Bar_altarar_fornecedor

        if pg == 'Cadastrar':
            tabela = self.ui.tabela_cadastrar_fornecedores
            produto = self.ui.line_search_Bar_cadastrar_fornecedores

        items = tabela.findItems(produto.text(), Qt.MatchContains)

        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    def SearchColaboradores(self, pg):
        tabela = self
        produto = self

        if pg == 'Colaboradores':
            tabela = self.ui.tabela_colaboradores
            produto = self.ui.line_search_bar_colaboradores

        if pg == 'Alterar':
            tabela = self.ui.tabela_alterar_colaboradores
            produto = self.ui.line_search_bar_buscar_alterar_colaboradores

        items = tabela.findItems(produto.text(), Qt.MatchContains)

        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    def SearchClientes(self, pg):
        tabela = self
        produto = self

        if pg == 'Clientes':
            tabela = self.ui.tabela_clientes
            produto = self.ui.line_search_Bar_clientes

        if pg == 'Alterar':
            tabela = self.ui.tabela_alterar_clientes
            produto = self.ui.line_search_Bar_alterar_clientes

        if pg == 'Cadastrar':
            tabela = self.ui.tabela_cadastrar_clientes
            produto = self.ui.line_search_Bar_cadastrar_clientes

        items = tabela.findItems(produto.text(), Qt.MatchContains)

        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    def SearchMonitoramentoVendas(self):
        tabela = self.ui.tabela_monitoramento_vendas
        vendas = self.ui.line_search_bar_monitoramentoto_vendas_vendas_vendas

        items = tabela.findItems(vendas.text(), Qt.MatchContains)
        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    def SearchMonitoramentoCompras(self):
        tabela = self.ui.tabela_monitoramento_compras
        vendas = self.ui.line_search_bar_monitoramentoto_vendas_vendas_compras

        items = tabela.findItems(vendas.text(), Qt.MatchContains)
        if items:
            item = items[0]
            tabela.setCurrentItem(item)

    # Funções para Atualiazar as previções das barras de pesquisa
    def AtualizaCompleterSearchFornecedores(self):
        global search_fornecedores

        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedores = cursor.fetchall()

        search_fornecedores.clear()
        search_fornecedores = []

        for fornecedor in banco_fornecedores:
            search_fornecedores.append(fornecedor[0])

        self.completer = QCompleter(search_fornecedores)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.ui.line_search_Bar_fornecedores.setCompleter(self.completer)
        self.ui.line_search_Bar_altarar_fornecedor.setCompleter(self.completer)
        self.ui.line_search_Bar_cadastrar_fornecedores.setCompleter(self.completer)
        self.ui.line_fornecedor_cadastrar_produto.setCompleter(self.completer)
        self.ui.line_fornecedor_alterar_produto.setCompleter(self.completer)

    def AtualizaCompleterSearchVendas(self):
        global search_monitoramento

        cursor.execute('SELECT * FROM monitoramento_vendas')
        banco_monitoramento = cursor.fetchall()

        search_monitoramento.clear()

        for venda in banco_monitoramento:
            if venda[0] not in search_monitoramento:
                search_monitoramento.append(venda[0])

            if venda[1] not in search_monitoramento:
                if venda[1] != 'Não Informado':
                    search_monitoramento.append(venda[1])
            search_monitoramento.append(venda[4])

        self.completer = QCompleter(search_monitoramento)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.ui.line_search_bar_monitoramentoto_vendas_vendas_vendas.setCompleter(self.completer)

    def AtualizaCompleterSearchCompras(self):
        global search_monitoramento_compras

        cursor.execute('SELECT * FROM monitoramento_compras')
        banco_monitoramento = cursor.fetchall()

        search_monitoramento_compras.clear()

        for compra in banco_monitoramento:
            if compra[0] not in search_monitoramento_compras:
                search_monitoramento.append(compra[0])

            if compra[1] not in search_monitoramento_compras:
                if compra[1] != 'Não Informado':
                    search_monitoramento.append(compra[1])
            search_monitoramento.append(compra[4])

        self.completer = QCompleter(search_monitoramento_compras)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.ui.line_search_bar_monitoramentoto_vendas_vendas_compras.setCompleter(self.completer)

    def AtualizaCompleterSearchProdutos(self):
        global search_produtos

        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        search_produtos.clear()

        for produto in banco_produtos:
            search_produtos.append(produto[1])

        self.completer = QCompleter(search_produtos)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)

        self.ui.line_search_Bar_produtos.setCompleter(self.completer)
        self.ui.line_search_Bar_alterar_produto.setCompleter(self.completer)
        self.ui.line_search_Bar_cadastrar_produto.setCompleter(self.completer)
        self.ui.line_search_bar_vendas.setCompleter(self.completer)
        self.ui.line_search_bar_vendas_carregamento.setCompleter(self.completer)

    def AtualizaCompleterSearchColaboradores(self):
        global search_colaboradores

        cursor.execute('SELECT * FROM login')
        banco_colaboradores = cursor.fetchall()

        search_colaboradores.clear()

        for colaborador in banco_colaboradores:
            search_colaboradores.append(colaborador[0])

        self.completer = QCompleter(search_colaboradores)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)

        self.ui.line_search_bar_buscar_alterar_colaboradores.setCompleter(self.completer)
        self.ui.line_search_bar_colaboradores.setCompleter(self.completer)

    def AtualizaCompleterSearchClientes(self):
        global search_clientes

        cursor.execute('SELECT * FROM clientes')
        banco_clientes = cursor.fetchall()

        search_clientes.clear()

        for clientes in banco_clientes:
            search_clientes.append(clientes[0])
            search_clientes.append(clientes[1])

        self.completer = QCompleter(search_clientes)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)

        self.ui.line_search_Bar_clientes.setCompleter(self.completer)
        self.ui.line_search_Bar_alterar_clientes.setCompleter(self.completer)
        self.ui.line_search_Bar_cadastrar_clientes.setCompleter(self.completer)
        self.ui.line_cliente.setCompleter(self.completer)

    # Funções de Cadastro
    def CadastroColaboradores(self):

        login = self.ui.line_login_cadastro_colaboradores
        senha = self.ui.line_senha_cadastro_colaboradores
        nome = self.ui.line_nome_cadastro_colaboradores
        nivel = ''


        if login.text() != '' and senha.text() != '' and nome.text() != '':
            nivel = 'admin'

            cursor.execute('SELECT * FROM login')
            banco_login = cursor.fetchall()

            LoginNoBanco = False

            for loginBanco in banco_login:

                if loginBanco[0] == login.text():
                    LoginNoBanco = True
                    break

            if LoginNoBanco == False:
                comando_SQL = 'INSERT INTO login VALUES (?,?,?,?)'
                dados = f'{login.text()}', f'{senha.text()}', f'{nivel}', f'{nome.text()}'
                cursor.execute(comando_SQL, dados)
                banco.commit()

                login.clear()
                senha.clear()
                nome.clear()

                self.ui.line_login_cadastro_colaboradores.setStyleSheet(StyleNormal)

                self.Popup('Cadastro de colaborador', 'Cadastro efetuado com sucesso')

            elif LoginNoBanco == True:
                self.Popup('Cadastro de colaborador', 'Este login ja existe')
                self.ui.line_login_cadastro_colaboradores.setStyleSheet(StyleError)

        self.AtualizaTabelasLogin()
        self.AtualizaCompleterSearchColaboradores()

    def CadastrarClientes(self):

        cpf = self.ui.line_cpf_cadastrar_clientes
        nome = self.ui.line_nome_cadastrar_clientes
        endereco = self.ui.line_endereco_cadastrar_clientes
        contato = self.ui.line_contato_cadastrar_clientes
        saldo_devedor = 'R$ 0,00'

        cursor.execute('SELECT nome FROM clientes WHERE nome = ?',(nome.text(),))
        vefNome = cursor.fetchall()
        cursor.execute('SELECT cpf FROM clientes WHERE cpf = ?', (cpf.text(),))
        vefCPF = cursor.fetchall()

        if cpf.text() != '' and nome.text() != '' and endereco.text() != '' and contato.text() != '':
             if not vefNome and not vefCPF:
                comando_SQL = 'INSERT INTO clientes (CPF, Nome, Endereço, Contato, saldo_devedor) VALUES (?,?,?,?,?)'
                dados = f'{cpf.text()}', f'{nome.text()}', f'{endereco.text()}', f'{contato.text()}', f'{saldo_devedor}'
                cursor.execute(comando_SQL, dados)
                banco.commit()

                self.AtualizaTabelasClientes()
                self.AtualizaCompleterSearchClientes()

                cpf.clear()
                nome.clear()
                endereco.clear()
                contato.clear()
             else:
                 self.Popup("Cadastrar cliente", "Este cliente ja esta cadastrado")
        else:
            self.Popup("Cadastrar cliente", "Insira os dados que faltam")
    def CadastrarFornecedores(self):

        nome = self.ui.line_cadastrar_nome_fornecedores
        endereco = self.ui.line_cadastrar_endereco_fornecedores
        contato = self.ui.line_cadastrar_contato_fornecedores

        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedores = cursor.fetchall()

        FornecedorNoBanco = False

        for fornecedor in banco_fornecedores:
            if fornecedor[0] == nome.text():
                FornecedorNoBanco = True

        if nome.text() != '' and endereco != '' and contato != '':
            if FornecedorNoBanco == False:
                comando_SQL = 'INSERT INTO fornecedores VALUES (?,?,?)'
                dados = f'{nome.text()}', f'{endereco.text()}', f'{contato.text()}'
                cursor.execute(comando_SQL, dados)
                banco.commit()

                nome.clear()
                endereco.clear()
                contato.clear()

                nome.setStyleSheet(StyleNormal)

                self.AtualizaTabelasFornecedores()
                self.AtualizaCompleterSearchFornecedores()
            else:
                nome.setStyleSheet(StyleError)

    def CadastrarProdutos(self):
        global search_fornecedores

        # Obtenha o maior código de produto atual
        cursor.execute('SELECT MAX(cód_produto) FROM produtos')
        maior_codigo = cursor.fetchone()

        # Se não houver produtos cadastrados, o maior_codigo será None, então comece com 1
        if maior_codigo[0] is None:
            cod_produto = 1
        else:
            cod_produto = int(maior_codigo[0]) + 1

        descricao = self.ui.line_descricao_cadastrar_produto.text()
        valor_unitario = self.ui.line_valor_cadastrar_produto_produto.text()
        qtde_estoque = self.ui.line_qtde_cadastrar_produto.text()
        fornecedor = self.ui.line_fornecedor_cadastrar_produto.text()
        valor_de_custo = self.ui.line_valorcusto_cadastrar_produto.text()
        valor_de_custo_verf = ''.join(re.findall(r'\d', valor_de_custo))

        cursor.execute("SELECT * FROM produtos")
        banco_produtos = cursor.fetchall()

        ProdutoJaCadastrado = False
        FornecedorNoSearch = False

        # Verifique se todos os campos necessários estão preenchidos corretamente
        if descricao != '' and valor_unitario != '' and qtde_estoque != '' and fornecedor != '' and valor_de_custo != '' and valor_de_custo_verf.isnumeric():
            for produto in banco_produtos:
                if produto[0] == str(cod_produto):
                    self.Popup("Erro - Cadastrar produto", "Código do produto já cadastrado")
                    ProdutoJaCadastrado = True
                    break

            if fornecedor in search_fornecedores:
                FornecedorNoSearch = True
                self.ui.line_fornecedor_cadastrar_produto.setStyleSheet(StyleNormal)
            else:
                self.Popup("Erro - Cadastrar produto", "Fornecedor não encontrado")

            if not ProdutoJaCadastrado and FornecedorNoSearch:
                comando_SQL = 'INSERT INTO produtos VALUES (?, ?, ?, ?, ?, ?)'
                dados = (str(cod_produto), descricao, valor_unitario, qtde_estoque, fornecedor, valor_de_custo)
                cursor.execute(comando_SQL, dados)
                banco.commit()

                # Limpa os campos da interface do usuário
                self.ui.line_descricao_cadastrar_produto.clear()
                self.ui.line_valor_cadastrar_produto_produto.clear()
                self.ui.line_qtde_cadastrar_produto.clear()
                self.ui.line_fornecedor_cadastrar_produto.clear()
                self.ui.line_valorcusto_cadastrar_produto.clear()

                self.AtualizaTabelasProdutos()
                self.AtualizaCompleterSearchProdutos()
        else:
            self.Popup('Erro - Cadastrar produto', 'Preencha todos os campos corretamente')

    def CadastrandoVendas(self):
        print("CadastrandoVendas iniciado")  # Adicionado para depuração
        global search_produtos, StyleError, StyleNormal

        produtoInserido = self.ui.line_codigo_vendas
        qtde = self.ui.line_quantidade_vendas
        desconto = self.ui.line_desconto_vendas
        NomeProduto = ''
        cliente = self.ui.line_cliente.text()

        ProdutoNoBanco = False
        QuantidadeMenorQueEstoque = False
        DescontoOk = False
        ValorUnitario = 0
        ValorDeCusto = 0

        print(f"Produto inserido: {produtoInserido.text()}")  # Adicionado para depuração
        print(f"Quantidade: {qtde.text()}")  # Adicionado para depuração
        print(f"Desconto: {desconto.text()}")  # Adicionado para depuração
        print(f"Cliente: {cliente}")  # Adicionado para depuração

        cursor.execute("SELECT * FROM produtos WHERE cód_produto = ?", (produtoInserido.text(),))
        banco_produtos = cursor.fetchall()
        print(f"Produtos no banco: {banco_produtos}")  # Adicionado para depuração

        desconto = ''.join(re.findall(r'\d', desconto.text()))

        if produtoInserido.text() != '' and produtoInserido.text().isnumeric() and cliente != '' and desconto.isnumeric():
            if banco_produtos:
                produto = banco_produtos[0]
                ProdutoNoBanco = True
                produtoInserido.setStyleSheet(StyleNormal)
                if qtde.text().isnumeric():
                    if (int(produto[3]) >= int(qtde.text()) or int(produto[3]) <= int(qtde.text())) and int(qtde.text()) > 0:
                        QuantidadeMenorQueEstoque = True
                        qtde.setStyleSheet(StyleNormal)

                        ValorUnitario = produto[2]
                        ValorDeCusto = produto[5]
                        NomeProduto = produto[1]
                        TotalQtde = int(produto[3]) - int(qtde.text())

                        valor = float(desconto) / 100 if desconto != '100' else 1.0
                        ValorUnitario = ''.join(re.findall(r'\d', ValorUnitario))
                        ValorDeCusto = ''.join(re.findall(r'\d', ValorDeCusto))
                        valorTotal = int(ValorUnitario) * int(qtde.text())
                        ValorDeCustoTotal = int(ValorDeCusto) * int(qtde.text())
                        descontoTotal = valorTotal * valor

                        if desconto.isnumeric() and descontoTotal <= valorTotal:
                            DescontoOk = True
                        else:
                            self.Popup("Erro - Adicionar compra", "Desconto não pode ser maior que o valor da compra")
                    else:
                        self.Popup("Erro - Adicionar compra", "Quantidade insuficiente no estoque")
                else:
                    self.Popup("Erro - Adicionar compra", "Informe um número para quantidade")
            else:
                self.Popup("Erro - Adicionar compra", "Código do produto é inexistente")

            if ProdutoNoBanco and QuantidadeMenorQueEstoque and DescontoOk:
                cursor.execute('SELECT MAX(id) FROM vendas')
                ultimo_id = cursor.fetchone()

                id = (ultimo_id[0] if ultimo_id[0] is not None else 0) + 1

                valorFinal = (valorTotal - descontoTotal) / 100
                valorFinal = format(valorFinal, '.2f')
                comando_SQL = 'INSERT INTO vendas VALUES (?,?,?,?,?,?,?)'
                dados = (produtoInserido.text(), NomeProduto, ValorUnitario, qtde.text(), valorFinal, id, ValorDeCustoTotal)
                print(f"Dados para inserção: {dados}")  # Adicionado para depuração
                cursor.execute(comando_SQL, dados)
                banco.commit()

            self.AtualizaTotal()
            self.AtualizaTabelasProdutos()
            self.AtualizaTabelaVendas()
        else:
            self.Popup("Erro - Adicionar compra", "Preencha os campos solicitados corretamente")

    def CadastrandoCompras(self):

        global search_produtos, StyleError, StyleNormal

        produtoInserido = self.ui.line_codigo_vendas_carregamento
        qtde = self.ui.line_quantidade_vendas_carregamento

        cursor.execute("SELECT descrição FROM produtos WHERE cód_produto = ?", (produtoInserido.text(),))
        nomeProduto = cursor.fetchone()

        cursor.execute("SELECT fornecedor FROM produtos WHERE cód_produto = ?", (produtoInserido.text(),))
        fornecedor = cursor.fetchone()

        QuantidadeMaiorQueZero = False
        ProdutoNoBanco = False
        ValorDeCusto = 0

        cursor.execute("SELECT * FROM produtos WHERE cód_produto = ?", (produtoInserido.text(),))
        banco_produtos = cursor.fetchall()

        if produtoInserido.text() != '' and produtoInserido.text().isnumeric() and qtde.text() != '' and qtde.text().isnumeric():
            if banco_produtos:
                produto = banco_produtos[0]
                ProdutoNoBanco = True
                produtoInserido.setStyleSheet(StyleNormal)
                if qtde.text().isnumeric() == True:
                    if int(qtde.text()) > 0:
                        QuantidadeMaiorQueZero = True
                        qtde.setStyleSheet(StyleNormal)

                        ValorDeCusto = float(produto[5].replace(',', '.'))
                        NomeProduto = produto[1]
                        TotalQtde = int(produto[3]) + int(qtde.text())

                        valorTotal = ValorDeCusto * int(qtde.text())

                    else:
                        self.Popup("Erro - Adicionar compra", "Preencha o campo de quantidade corretamente")
                else:
                    self.Popup("Erro - Adicionar compra", "Informe um numero para quantidade")
            else:
                self.Popup("Erro - Adicionar compra", "Código do produto é inexistente")

            if ProdutoNoBanco == True and QuantidadeMaiorQueZero == True:
                cursor.execute('SELECT MAX(id) FROM compras')
                ultimo_id = cursor.fetchone()

                if ultimo_id[0] is None:
                    id = 0
                else:
                    id = int(ultimo_id[0]) + 1

                valorFinal = format(valorTotal, '.2f')
                comando_SQL = 'INSERT INTO compras VALUES (?,?,?,?,?,?,?)'
                dados = (
                    produtoInserido.text(),
                    nomeProduto[0],  # Extrair o nome do produto da tupla
                    ValorDeCusto,
                    qtde.text(),
                    valorFinal,
                    id,
                    fornecedor[0]  # Extrair o fornecedor da tupla
                )
                cursor.execute(comando_SQL, dados)
                banco.commit()

            self.AtualizaTotalCompras()
            self.AtualizaTabelasProdutos()
            self.AtualizaTabelaCompras()
        else:
            self.Popup("Erro - Adicionar compra", "Preencha os campos solicitados corretamente")

    # Funções de Alterar
    def AlterarColaboradores(self):
        global id_tabela_alterar

        login = self.ui.line_login_alterar_colaboradores
        senha = self.ui.line_senha_alterar_colaboradores
        nome = self.ui.line_nome_alterar_colaboradores

        cursor.execute('SELECT * FROM login')
        banco_login = cursor.fetchall()

        if login.text() != '' and senha.text() != '' and nome.text() != '':

            LoginNoBanco = False

            for pos, user in enumerate(banco_login):
                if login.text() == user[0] and pos != id_tabela_alterar:
                    LoginNoBanco = True

            for pos, user in enumerate(banco_login):
                if pos == id_tabela_alterar:
                    if LoginNoBanco == False:
                        cursor.execute(
                            f'UPDATE login set usuario = "{login.text()}", senha = "{senha.text()}", nivel = "{user[2]}", nome = "{nome.text()}"'
                            f'WHERE usuario = "{user[0]}"')
                        banco.commit()

                        login.clear()
                        senha.clear()
                        nome.clear()

                        self.AtualizaTabelasLogin()
                        self.AtualizaCompleterSearchColaboradores()

                        self.ui.line_login_alterar_colaboradores.setStyleSheet(StyleNormal)
                        break
                    else:
                        self.ui.line_login_alterar_colaboradores.setStyleSheet(StyleError)

    def AlterarClientes(self):
        global id_alterar_Clientes

        cpf = self.ui.line_alterar_cpf_alterar_clientes
        nome = self.ui.line_alterar_nome_alterar_clientes
        endereco = self.ui.line_alterar_endereco_alterar_clientes
        contato = self.ui.line_alterar_contato_alterar_clientes

        cursor.execute('SELECT * FROM clientes')
        banco_clientes = cursor.fetchall()
        if cpf.text() != '' and nome.text() != '' and endereco.text() != '' and contato.text() != '':
            for pos, cliente in enumerate(banco_clientes):
                if pos == id_alterar_Clientes:
                    cursor.execute(
                        f'UPDATE clientes set CPF = "{cpf.text()}", nome = "{nome.text()}", endereço = "{endereco.text()}", contato = "{contato.text()}"'
                        f'WHERE CPF = "{cliente[0]}"')

                    cpf.clear()
                    nome.clear()
                    endereco.clear()
                    contato.clear()

                    self.AtualizaTabelasClientes()
                    self.AtualizaCompleterSearchClientes()

                    break

    def AlterarFornecedores(self):
        global id_alterar_fornecedores

        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedores = cursor.fetchall()

        nome = self.ui.line_alterar_nome_fornecedor
        endereco = self.ui.line_alterar_endereco_fornecedor
        contato = self.ui.line_alterar_contato_fornecedor

        if nome.text() != '' and endereco.text() != '' and contato.text() != '':
            for pos, fornecedores in enumerate(banco_fornecedores):
                if pos == id_alterar_fornecedores:
                    cursor.execute(
                        f'UPDATE fornecedores set nome = "{nome.text()}", endereço = "{endereco.text()}", contato = "{contato.text()}"'
                        f'WHERE nome = "{fornecedores[0]}"')

                    nome.clear()
                    endereco.clear()
                    contato.clear()

                    self.AtualizaTabelasFornecedores()
                    self.AtualizaCompleterSearchFornecedores()
                    break

    def AlterarProdutos(self):
        global id_alterar_produtos
        global search_fornecedores

        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        cod_produto = self.ui.line_codigo_alterar_produto
        descricao = self.ui.line_decricao_alterar_produto
        valor_unitario = self.ui.line_valor_alterar_produto
        valor_de_custo = self.ui.line_valorcusto_alterar_produto
        qtde_estoque = self.ui.line_qtde_alterar_produto
        fornecedor = self.ui.line_fornecedor_alterar_produto

        FornecedorNoSearch = False
        ProdutoJaCadastrado = False
        AlterarProduto = ''

        if cod_produto.text() != '' and descricao.text() != '' and valor_unitario.text() != '' and qtde_estoque.text() != '' and fornecedor.text() != '' and valor_de_custo.text() != '':
            if fornecedor.text() in search_fornecedores:
                FornecedorNoSearch = True

                fornecedor.setStyleSheet(StyleNormal)

            else:
                fornecedor.setStyleSheet(StyleError)
                self.Popup("Erro - Alterar produtos", "Fornecedor não encontrado")

            for pos, produto in enumerate(banco_produtos):
                if cod_produto.text() == produto[0] and pos != id_alterar_produtos:
                    ProdutoJaCadastrado = True

                    self.Popup("Erro - Alterar produtos", "Produto ja cadastrado")

                else:
                    cod_produto.setStyleSheet(StyleNormal)

                if pos == id_alterar_produtos:
                    AlterarProduto = produto[0]
        else:
            self.Popup("Erro - Alterar produtos", "Preencha todos campos corretamente")

        if FornecedorNoSearch == True and ProdutoJaCadastrado == False:
            cursor.execute(
                f'UPDATE produtos set cód_produto = ?, descrição = ?, valor_unitário = ?, qtde_estoque = ?, fornecedor = ?, valor_de_custo = ? WHERE cód_produto = ?',
                (
                    cod_produto.text(),
                    descricao.text(),
                    valor_unitario.text(),
                    qtde_estoque.text(),
                    fornecedor.text(),
                    valor_de_custo.text(),
                    AlterarProduto,)
            )
            banco.commit()

            cod_produto.clear()
            descricao.clear()
            valor_unitario.clear()
            qtde_estoque.clear()
            fornecedor.clear()
            valor_de_custo.clear()

            self.AtualizaTabelasProdutos()
            self.AtualizaCompleterSearchProdutos()
            self.AtualizaTabelaVendas()

    # Funções de Excluir
    def ExcluirColaboradores(self):

        id = self.ui.tabela_colaboradores.currentRow()

        cursor.execute('SELECT * FROM login')
        banco_login = cursor.fetchall()

        deletar_user = ''

        for pos, user in enumerate(banco_login):
            if id == pos:
                deletar_user = user[0]

        cursor.execute(f'DELETE FROM login WHERE usuario = "{deletar_user}"')
        banco.commit()

        self.AtualizaTabelasLogin()
        self.AtualizaCompleterSearchColaboradores()

    def ExcluirClientes(self):
        id = self.ui.tabela_clientes.currentRow()

        cursor.execute('SELECT * FROM clientes')
        banco_clientes = cursor.fetchall()

        deletar_cliente = ''

        for pos, cliente in enumerate(banco_clientes):
            if id == pos:
                deletar_cliente = cliente[0]

        cursor.execute(f'DELETE FROM clientes WHERE CPF = "{deletar_cliente}"')
        banco.commit()

        self.AtualizaTabelasClientes()
        self.AtualizaCompleterSearchClientes()

    def ExluirFornecedores(self):
        id = self.ui.tabela_fornecedores.currentRow()

        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedor = cursor.fetchall()

        deletar_fornecedor = ''

        for pos, fornecedor in enumerate(banco_fornecedor):
            if id == pos:
                deletar_fornecedor = fornecedor[0]

        cursor.execute(f'DELETE FROM fornecedores WHERE Nome = "{deletar_fornecedor}"')
        banco.commit()

        self.AtualizaTabelasFornecedores()
        self.AtualizaCompleterSearchFornecedores()

    def ExcluirProdutos(self):

        id = self.ui.tabela_produto.currentRow()

        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        for pos, produto in enumerate(banco_produtos):
            if pos == id:
                cursor.execute(f'DELETE FROM produtos WHERE cód_produto = "{produto[0]}"')
                banco.commit()

                self.AtualizaTabelasProdutos()
                self.AtualizaCompleterSearchProdutos()

    def ExcluirVenda(self):
        id = self.ui.tabela_vendas.currentRow()

        if id != -1:
            cursor.execute('SELECT * FROM vendas ORDER BY id ASC')
            banco_vendas = cursor.fetchall()
            cursor.execute('SELECT * FROM produtos')
            banco_produtos = cursor.fetchall()

            if id < len(banco_vendas):
                venda_selecionada = banco_vendas[id]
                id_deletado = venda_selecionada[5]

                for produto in banco_produtos:
                    if venda_selecionada[0] == produto[0]:
                        TotalEstoque = int(venda_selecionada[3]) + int(produto[3])
                        cursor.execute('UPDATE produtos SET qtde_estoque = ? WHERE cód_produto = ?',
                                       (TotalEstoque, produto[0]))
                        break

                cursor.execute('DELETE FROM vendas WHERE id = ?', (id_deletado,))
                banco.commit()

                cursor.execute('SELECT * FROM vendas ORDER BY id ASC')
                banco_vendas = cursor.fetchall()

                # Reordenar os IDs para começar em 1
                for index, venda in enumerate(banco_vendas, start=1):
                    if venda[5] != index:
                        cursor.execute('UPDATE vendas SET id = ? WHERE id = ?', (index, venda[5]))
                        banco.commit()

            self.AtualizaTabelaVendas()
            self.AtualizaTotal()
            self.AtualizaTabelasProdutos()

    def ExcluirCompra(self):
        id = self.ui.tabela_vendas_carregamento.currentRow()

        if id != -1:
            cursor.execute('SELECT * FROM compras ORDER BY id ASC')
            banco_compras = cursor.fetchall()
            cursor.execute('SELECT * FROM produtos')
            banco_produtos = cursor.fetchall()

            if id < len(banco_compras):
                compra_selecionada = banco_compras[id]
                id_deletado = compra_selecionada[5]

                cursor.execute('DELETE FROM compras WHERE id = ?', (id_deletado,))
                banco.commit()

                cursor.execute('SELECT * FROM vendas ORDER BY id ASC')
                banco_compras = cursor.fetchall()

                # Reordenar os IDs para começar em 1
                for index, compra in enumerate(banco_compras, start=1):
                    if compra[5] != index:
                        cursor.execute('UPDATE compras SET id = ? WHERE id = ?', (index, compra[5]))
                        banco.commit()

            self.AtualizaTabelaCompras()
            self.AtualizaTotalCompras()
            self.AtualizaTabelasProdutos()

    # Funções de setar Texto
    def setTextAlterarColaboradores(self):
        global id_tabela_alterar
        nome = self.ui.line_nome_alterar_colaboradores
        login = self.ui.line_login_alterar_colaboradores
        senha = self.ui.line_senha_alterar_colaboradores

        id_tabela_alterar = self.ui.tabela_alterar_colaboradores.currentRow()

        cursor.execute('SELECT * FROM login')
        banco_login = cursor.fetchall()

        for pos, user in enumerate(banco_login):
            if pos == id_tabela_alterar:
                nome.setText(user[3])
                login.setText(user[0])
                senha.setText(user[1])

    def setTextAlterarClientes(self):
        global id_alterar_Clientes
        cpf = self.ui.line_alterar_cpf_alterar_clientes
        nome = self.ui.line_alterar_nome_alterar_clientes
        endereco = self.ui.line_alterar_endereco_alterar_clientes
        contato = self.ui.line_alterar_contato_alterar_clientes

        id_alterar_Clientes = self.ui.tabela_alterar_clientes.currentRow()

        cursor.execute('SELECT * FROM clientes')
        banco_clientes = cursor.fetchall()

        for pos, cliente in enumerate(banco_clientes):
            if pos == id_alterar_Clientes:
                cpf.setText(cliente[0])
                nome.setText(cliente[1])
                endereco.setText(cliente[2])
                contato.setText(cliente[3])

    def setTextAlterarFornecedores(self):
        global id_alterar_fornecedores

        nome = self.ui.line_alterar_nome_fornecedor
        endereco = self.ui.line_alterar_endereco_fornecedor
        contato = self.ui.line_alterar_contato_fornecedor

        id_alterar_fornecedores = self.ui.tabela_alterar_fornecedores.currentRow()

        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedores = cursor.fetchall()

        for pos, fornecedor in enumerate(banco_fornecedores):
            if pos == id_alterar_fornecedores:
                nome.setText(fornecedor[0])
                endereco.setText(fornecedor[1])
                contato.setText(fornecedor[2])

    def setTextAlterarProdutos(self):
        global id_alterar_produtos

        cod_produto = self.ui.line_codigo_alterar_produto
        descricao = self.ui.line_decricao_alterar_produto
        valor_unitario = self.ui.line_valor_alterar_produto
        qtde_estoque = self.ui.line_qtde_alterar_produto
        fornecedor = self.ui.line_fornecedor_alterar_produto
        valor_de_custo = self.ui.line_valorcusto_alterar_produto

        id_alterar_produtos = self.ui.tabela_alterar_produto.currentRow()

        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        for pos, produto in enumerate(banco_produtos):
            if pos == id_alterar_produtos:
                cod_produto.setText(produto[0])
                descricao.setText(produto[1])
                valor_unitario.setText(produto[2])
                qtde_estoque.setText(produto[3])
                fornecedor.setText(produto[4])
                valor_de_custo.setText(produto[5])

    # Funções para ver Senha
    def VerSenhaCadastroSangria(self):
        global click_cadastro_sangria

        click_cadastro_sangria += 1

        if click_cadastro_sangria % 2 == 0:
            self.ui.line_senha_gerar_sangria.setEchoMode(QLineEdit.EchoMode.Password)
            self.ui.btn_ver_senha_gerar_sangria.setStyleSheet('QPushButton {'
                                                'background-image: url(:/icones/ver senha.png);'
                                                'border: 0px;'
                                                'outline: 0;'
                                                '}'
                                                ''
                                                'QPushButton:hover {'
                                                'background-image: url(:/icones/ver senha hover.png);'
                                                '}')

        if click_cadastro_sangria % 2 == 1:
            self.ui.line_senha_gerar_sangria.setEchoMode(QLineEdit.EchoMode.Normal)
            self.ui.btn_ver_senha_gerar_sangria.setStyleSheet('QPushButton {'
                                                'background-image: url(:/icones/bloquear senha.png);'
                                                'border: 0px;'
                                                'outline: 0;'
                                                '}'
                                                ''
                                                'QPushButton:hover {'
                                                'background-image: url(:/icones/bloquear senha hover.png);''}')


    def VerSenhaCadastroColaboradores(self):
        global click_cadastro_colaboradores

        click_cadastro_colaboradores += 1

        if click_cadastro_colaboradores % 2 == 0:
            self.ui.line_senha_cadastro_colaboradores.setEchoMode(QLineEdit.EchoMode.Password)
            self.ui.btn_ver_senha_cadastro_colaboradores.setStyleSheet('QPushButton {'
                                                'background-image: url(:/icones/ver senha.png);'
                                                'border: 0px;'
                                                'outline: 0;'
                                                '}'
                                                ''
                                                'QPushButton:hover {'
                                                'background-image: url(:/icones/ver senha hover.png);'
                                                '}')

        if click_cadastro_colaboradores % 2 == 1:
            self.ui.line_senha_cadastro_colaboradores.setEchoMode(QLineEdit.EchoMode.Normal)
            self.ui.btn_ver_senha_cadastro_colaboradores.setStyleSheet('QPushButton {'
                                                'background-image: url(:/icones/bloquear senha.png);'
                                                'border: 0px;'
                                                'outline: 0;'
                                                '}'
                                                ''
                                                'QPushButton:hover {'
                                                'background-image: url(:/icones/bloquear senha hover.png);''}')

    def VerSenhaAlterarColaboradores(self):
        global click_alterar_colaboradores

        click_alterar_colaboradores += 1

        if click_alterar_colaboradores % 2 == 0:
            self.ui.line_senha_alterar_colaboradores.setEchoMode(QLineEdit.EchoMode.Password)
            self.ui.btn_ver_senha_alterar_colaboradores.setStyleSheet('QPushButton {'
                                                        'background-image: url(:/icones/ver senha.png);'
                                                        'border: 0px;'
                                                        'outline: 0;'
                                                        '}'
                                                        ''
                                                        'QPushButton:hover {'
                                                        'background-image: url(:/icones/ver senha hover.png);'
                                                        '}')
        if click_alterar_colaboradores % 2 == 1:
            self.ui.line_senha_alterar_colaboradores.setEchoMode(QLineEdit.EchoMode.Normal)
            self.ui.btn_ver_senha_alterar_colaboradores.setStyleSheet('QPushButton {'
                                                        'background-image: url(:/icones/bloquear senha.png);'
                                                        'border: 0px;'
                                                        'outline: 0;'
                                                        '}'
                                                        ''
                                                        'QPushButton:hover {'
                                                        'background-image: url(:/icones/bloquear senha hover.png);''}')

    # Funções para Atualizar Tabelas
    def AtualizaTabelasLogin(self):
        cursor.execute('SELECT * FROM login')
        banco_login = cursor.fetchall()

        self.ui.tabela_colaboradores.clear()
        self.ui.tabela_alterar_colaboradores.clear()

        row = 0
        self.ui.tabela_colaboradores.setRowCount(len(banco_login))
        self.ui.tabela_alterar_colaboradores.setRowCount(len(banco_login))

        colunas = ['Nome', 'Login', 'Senha']
        self.ui.tabela_colaboradores.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_alterar_colaboradores.setHorizontalHeaderLabels(colunas)

        for pos, logins in enumerate(banco_login):
            self.ui.tabela_colaboradores.setItem(row, 0, QTableWidgetItem(logins[3]))
            self.ui.tabela_colaboradores.setItem(row, 1, QTableWidgetItem(logins[0]))
            self.ui.tabela_colaboradores.setItem(row, 2, QTableWidgetItem(logins[1]))

            self.ui.tabela_alterar_colaboradores.setItem(row, 0, QTableWidgetItem(logins[3]))
            self.ui.tabela_alterar_colaboradores.setItem(row, 1, QTableWidgetItem(logins[0]))
            self.ui.tabela_alterar_colaboradores.setItem(row, 2, QTableWidgetItem(logins[1]))

            row += 1

        # Definir uma altura fixa para todas as linhas
        row_count = self.ui.tabela_colaboradores.rowCount()
        for row in range(row_count):
            self.ui.tabela_colaboradores.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels
            self.ui.tabela_alterar_colaboradores.setRowHeight(row, 50)

        # Ajustar o tamanho da fonte para os itens das tabelas
        font = QtGui.QFont()
        font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
        self.ui.tabela_colaboradores.setFont(font)
        self.ui.tabela_alterar_colaboradores.setFont(font)

        # Habilitar a quebra de linha automática para todas as células das tabelas
        self.ui.tabela_colaboradores.resizeRowsToContents()
        self.ui.tabela_alterar_colaboradores.resizeRowsToContents()

        # Ajustar o tamanho da fonte para os cabeçalhos das colunas das tabelas
        header_font = QtGui.QFont()
        header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
        self.ui.tabela_colaboradores.horizontalHeader().setFont(header_font)
        self.ui.tabela_alterar_colaboradores.horizontalHeader().setFont(header_font)

        # Ajustar a altura dos cabeçalhos das colunas das tabelas
        self.ui.tabela_colaboradores.horizontalHeader().setFixedHeight(
            40)  # Ajuste a altura do cabeçalho conforme necessário
        self.ui.tabela_alterar_colaboradores.horizontalHeader().setFixedHeight(40)

    def AtualizaTabelasClientes(self):
        cursor.execute('SELECT * FROM clientes')
        banco_clientes = cursor.fetchall()

        self.ui.tabela_clientes.clear()
        self.ui.tabela_alterar_clientes.clear()
        self.ui.tabela_cadastrar_clientes.clear()

        row = 0

        self.ui.tabela_clientes.setRowCount(len(banco_clientes))
        self.ui.tabela_alterar_clientes.setRowCount(len(banco_clientes))
        self.ui.tabela_cadastrar_clientes.setRowCount(len(banco_clientes))

        colunas = ['CPF', 'Nome', 'Endereço', 'Contato', 'Saldo devedor']
        self.ui.tabela_clientes.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_alterar_clientes.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_cadastrar_clientes.setHorizontalHeaderLabels(colunas)

        for clientes in banco_clientes:
            self.ui.tabela_clientes.setItem(row, 0, QTableWidgetItem(clientes[0]))
            self.ui.tabela_clientes.setItem(row, 1, QTableWidgetItem(clientes[1]))
            self.ui.tabela_clientes.setItem(row, 2, QTableWidgetItem(clientes[2]))
            self.ui.tabela_clientes.setItem(row, 3, QTableWidgetItem(clientes[3]))
            self.ui.tabela_clientes.setItem(row, 4, QTableWidgetItem(clientes[4]))

            self.ui.tabela_alterar_clientes.setItem(row, 0, QTableWidgetItem(clientes[0]))
            self.ui.tabela_alterar_clientes.setItem(row, 1, QTableWidgetItem(clientes[1]))
            self.ui.tabela_alterar_clientes.setItem(row, 2, QTableWidgetItem(clientes[2]))
            self.ui.tabela_alterar_clientes.setItem(row, 3, QTableWidgetItem(clientes[3]))
            self.ui.tabela_alterar_clientes.setItem(row, 4, QTableWidgetItem(clientes[4]))

            self.ui.tabela_cadastrar_clientes.setItem(row, 0, QTableWidgetItem(clientes[0]))
            self.ui.tabela_cadastrar_clientes.setItem(row, 1, QTableWidgetItem(clientes[1]))
            self.ui.tabela_cadastrar_clientes.setItem(row, 2, QTableWidgetItem(clientes[2]))
            self.ui.tabela_cadastrar_clientes.setItem(row, 3, QTableWidgetItem(clientes[3]))
            self.ui.tabela_cadastrar_clientes.setItem(row, 4, QTableWidgetItem(clientes[4]))
            row += 1

        # Definir uma altura fixa para todas as linhas das tabelas
        for table in [self.ui.tabela_clientes, self.ui.tabela_alterar_clientes, self.ui.tabela_cadastrar_clientes]:
            row_count = table.rowCount()
            for row in range(row_count):
                table.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

            # Ajustar o tamanho da fonte para os itens das tabelas
            font = QtGui.QFont()
            font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
            table.setFont(font)

            # Habilitar a quebra de linha automática para todas as células das tabelas
            table.resizeRowsToContents()

            # Ajustar o tamanho da fonte para os cabeçalhos das colunas das tabelas
            header_font = QtGui.QFont()
            header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
            table.horizontalHeader().setFont(header_font)

            # Ajustar a altura dos cabeçalhos das colunas das tabelas
            table.horizontalHeader().setFixedHeight(40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelasFornecedores(self):
        cursor.execute('SELECT * FROM fornecedores')
        banco_fornecedores = cursor.fetchall()

        self.ui.tabela_fornecedores.clear()
        self.ui.tabela_cadastrar_fornecedores.clear()
        self.ui.tabela_alterar_fornecedores.clear()

        row = 0

        self.ui.tabela_fornecedores.setRowCount(len(banco_fornecedores))
        self.ui.tabela_cadastrar_fornecedores.setRowCount(len(banco_fornecedores))
        self.ui.tabela_alterar_fornecedores.setRowCount(len(banco_fornecedores))

        colunas = ['Nome', 'Endereço', 'Contato']
        self.ui.tabela_fornecedores.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_alterar_fornecedores.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_cadastrar_fornecedores.setHorizontalHeaderLabels(colunas)

        for fornecedores in banco_fornecedores:
            self.ui.tabela_fornecedores.setItem(row, 0, QTableWidgetItem(fornecedores[0]))
            self.ui.tabela_fornecedores.setItem(row, 1, QTableWidgetItem(fornecedores[1]))
            self.ui.tabela_fornecedores.setItem(row, 2, QTableWidgetItem(fornecedores[2]))

            self.ui.tabela_alterar_fornecedores.setItem(row, 0, QTableWidgetItem(fornecedores[0]))
            self.ui.tabela_alterar_fornecedores.setItem(row, 1, QTableWidgetItem(fornecedores[1]))
            self.ui.tabela_alterar_fornecedores.setItem(row, 2, QTableWidgetItem(fornecedores[2]))

            self.ui.tabela_cadastrar_fornecedores.setItem(row, 0, QTableWidgetItem(fornecedores[0]))
            self.ui.tabela_cadastrar_fornecedores.setItem(row, 1, QTableWidgetItem(fornecedores[1]))
            self.ui.tabela_cadastrar_fornecedores.setItem(row, 2, QTableWidgetItem(fornecedores[2]))
            row += 1

        # Definir uma altura fixa para todas as linhas das tabelas
        for table in [self.ui.tabela_fornecedores, self.ui.tabela_cadastrar_fornecedores,
                      self.ui.tabela_alterar_fornecedores]:
            row_count = table.rowCount()
            for row in range(row_count):
                table.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

            # Ajustar o tamanho da fonte para os itens das tabelas
            font = QtGui.QFont()
            font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
            table.setFont(font)

            # Habilitar a quebra de linha automática para todas as células das tabelas
            table.resizeRowsToContents()

            # Ajustar o tamanho da fonte para os cabeçalhos das colunas das tabelas
            header_font = QtGui.QFont()
            header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
            table.horizontalHeader().setFont(header_font)

            # Ajustar a altura dos cabeçalhos das colunas das tabelas
            table.horizontalHeader().setFixedHeight(40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelasProdutos(self):
        cursor.execute('SELECT * FROM produtos')
        banco_produtos = cursor.fetchall()

        self.ui.tabela_produto.clear()
        self.ui.tabela_alterar_produto.clear()
        self.ui.tabela_cadastro_produto.clear()

        row = 0

        self.ui.tabela_produto.setRowCount(len(banco_produtos))
        self.ui.tabela_alterar_produto.setRowCount(len(banco_produtos))
        self.ui.tabela_cadastro_produto.setRowCount(len(banco_produtos))

        colunas = ['Item', 'Cód', 'Produto', 'Valor De Venda', 'Qtde', 'Fornecedor', 'Valor de custo']
        self.ui.tabela_produto.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_alterar_produto.setHorizontalHeaderLabels(colunas)
        self.ui.tabela_cadastro_produto.setHorizontalHeaderLabels(colunas)

        for pos, produto in enumerate(banco_produtos):
            valor_unitario = produto[2]
            valor_de_custo = produto[5]

            for table in [self.ui.tabela_produto, self.ui.tabela_alterar_produto, self.ui.tabela_cadastro_produto]:
                table.setItem(row, 0, QTableWidgetItem(f'{pos + 1}'))
                table.setItem(row, 1, QTableWidgetItem(produto[0]))
                table.setItem(row, 2, QTableWidgetItem(produto[1]))
                table.setItem(row, 3, QTableWidgetItem('R$ ' + valor_unitario))
                table.setItem(row, 4, QTableWidgetItem(produto[3]))
                table.setItem(row, 5, QTableWidgetItem(produto[4]))
                table.setItem(row, 6, QTableWidgetItem('R$ ' + valor_de_custo))

            row += 1

        # Definir uma altura fixa para todas as linhas das tabelas
        for table in [self.ui.tabela_produto, self.ui.tabela_alterar_produto, self.ui.tabela_cadastro_produto]:
            row_count = table.rowCount()
            for row in range(row_count):
                table.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

            # Ajustar o tamanho da fonte para os itens das tabelas
            font = QtGui.QFont()
            font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
            table.setFont(font)

            # Habilitar a quebra de linha automática para todas as células das tabelas
            table.resizeRowsToContents()

            # Ajustar o tamanho da fonte para os cabeçalhos das colunas das tabelas
            header_font = QtGui.QFont()
            header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
            table.horizontalHeader().setFont(header_font)

            # Ajustar a altura dos cabeçalhos das colunas das tabelas
            table.horizontalHeader().setFixedHeight(40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelaMonitoramentoVendas(self):
        cursor.execute('SELECT * FROM monitoramento_vendas')
        banco_monitoramento = cursor.fetchall()

        self.ui.tabela_monitoramento_vendas.clear()

        row = 0

        self.ui.tabela_monitoramento_vendas.setRowCount(len(banco_monitoramento))

        colunas = ['Vendedor', 'Cliente', 'Produto', 'Qtde Vendido', 'Total Venda', 'Data/horário', 'A Prazo', 'Forma de Pagamento']
        self.ui.tabela_monitoramento_vendas.setHorizontalHeaderLabels(colunas)

        for venda in banco_monitoramento:

            if venda[6] == 1:
                se_prazo = 'Sim'
            else:
                se_prazo = 'Não'

            self.ui.tabela_monitoramento_vendas.setItem(row, 0, QTableWidgetItem(venda[0]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 1, QTableWidgetItem(venda[1]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 2, QTableWidgetItem(venda[7]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 3, QTableWidgetItem(venda[2]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 4, QTableWidgetItem('R$ ' + venda[3]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 5, QTableWidgetItem(venda[4]))
            self.ui.tabela_monitoramento_vendas.setItem(row, 6, QTableWidgetItem(se_prazo))
            self.ui.tabela_monitoramento_vendas.setItem(row, 7, QTableWidgetItem(venda[9]))

            row += 1

        # Definir uma altura fixa para todas as linhas da tabela
        row_count = self.ui.tabela_monitoramento_vendas.rowCount()
        for row in range(row_count):
            self.ui.tabela_monitoramento_vendas.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

        # Ajustar o tamanho da fonte para os itens da tabela
        font = QtGui.QFont()
        font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
        self.ui.tabela_monitoramento_vendas.setFont(font)

        # Habilitar a quebra de linha automática para todas as células da tabela
        self.ui.tabela_monitoramento_vendas.resizeRowsToContents()

        # Ajustar o tamanho da fonte para os cabeçalhos das colunas da tabela
        header_font = QtGui.QFont()
        header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
        self.ui.tabela_monitoramento_vendas.horizontalHeader().setFont(header_font)

        # Ajustar a altura do cabeçalho da coluna da tabela
        self.ui.tabela_monitoramento_vendas.horizontalHeader().setFixedHeight(
            40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelaMonitoramentoCompras(self):
        cursor.execute('SELECT * FROM monitoramento_compras')
        banco_monitoramento = cursor.fetchall()

        self.ui.tabela_monitoramento_compras.clear()

        row = 0

        self.ui.tabela_monitoramento_compras.setRowCount(len(banco_monitoramento))

        colunas = ['Comprador', 'Fornecedor', 'Produto', 'Qtde Comprado', 'Total Compra', 'Data/horário']
        self.ui.tabela_monitoramento_compras.setHorizontalHeaderLabels(colunas)

        for compra in banco_monitoramento:
            self.ui.tabela_monitoramento_compras.setItem(row, 0, QTableWidgetItem(compra[0]))
            self.ui.tabela_monitoramento_compras.setItem(row, 1, QTableWidgetItem(compra[1]))
            self.ui.tabela_monitoramento_compras.setItem(row, 2, QTableWidgetItem(compra[5]))
            self.ui.tabela_monitoramento_compras.setItem(row, 3, QTableWidgetItem(compra[2]))
            self.ui.tabela_monitoramento_compras.setItem(row, 4, QTableWidgetItem('R$ ' + compra[3]))
            self.ui.tabela_monitoramento_compras.setItem(row, 5, QTableWidgetItem(compra[4]))
            row += 1

        # Definir uma altura fixa para todas as linhas da tabela
        row_count = self.ui.tabela_monitoramento_compras.rowCount()
        for row in range(row_count):
            self.ui.tabela_monitoramento_compras.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

        # Ajustar o tamanho da fonte para os itens da tabela
        font = QtGui.QFont()
        font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
        self.ui.tabela_monitoramento_compras.setFont(font)

        # Habilitar a quebra de linha automática para todas as células da tabela
        self.ui.tabela_monitoramento_compras.resizeRowsToContents()

        # Ajustar o tamanho da fonte para os cabeçalhos das colunas da tabela
        header_font = QtGui.QFont()
        header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
        self.ui.tabela_monitoramento_compras.horizontalHeader().setFont(header_font)

        # Ajustar a altura do cabeçalho da coluna da tabela
        self.ui.tabela_monitoramento_compras.horizontalHeader().setFixedHeight(
            40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelaVendas(self):
        cursor.execute('SELECT * FROM vendas ORDER BY id ASC')
        banco_vendas = cursor.fetchall()

        row = 0

        self.ui.tabela_vendas.setRowCount(len(banco_vendas))
        self.ui.tabela_vendas.clear()

        colunas = ['Item', 'Cód', 'Produto', 'Valor De Venda', 'Qtde', 'Total']
        self.ui.tabela_vendas.setHorizontalHeaderLabels(colunas)

        for venda in banco_vendas:
            valor_unitario = ''.join(re.findall(r'\d', venda[2]))
            valor_unitario = lang.toString(int(valor_unitario) * 0.01, 'f', 2)
            total = venda[4]

            self.ui.tabela_vendas.setItem(row, 0, QTableWidgetItem(f'{venda[5]}'))
            self.ui.tabela_vendas.setItem(row, 1, QTableWidgetItem(venda[0]))
            self.ui.tabela_vendas.setItem(row, 2, QTableWidgetItem(venda[1]))
            self.ui.tabela_vendas.setItem(row, 3, QTableWidgetItem('R$ ' + valor_unitario))
            self.ui.tabela_vendas.setItem(row, 4, QTableWidgetItem(venda[3]))
            self.ui.tabela_vendas.setItem(row, 5, QTableWidgetItem('R$ ' + total))

            row += 1

        # Definir uma altura fixa para todas as linhas da tabela
        row_count = self.ui.tabela_vendas.rowCount()
        for row in range(row_count):
            self.ui.tabela_vendas.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

        # Ajustar o tamanho da fonte para os itens da tabela
        font = QtGui.QFont()
        font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
        self.ui.tabela_vendas.setFont(font)

        # Habilitar a quebra de linha automática para todas as células da tabela
        self.ui.tabela_vendas.resizeRowsToContents()

        # Ajustar o tamanho da fonte para os cabeçalhos das colunas da tabela
        header_font = QtGui.QFont()
        header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
        self.ui.tabela_vendas.horizontalHeader().setFont(header_font)

        # Ajustar a altura do cabeçalho da coluna da tabela
        self.ui.tabela_vendas.horizontalHeader().setFixedHeight(40)  # Ajuste a altura do cabeçalho conforme necessário

    def AtualizaTabelaCompras(self):
        cursor.execute('SELECT * FROM compras ORDER BY id ASC')
        banco_compras = cursor.fetchall()

        row = 0

        self.ui.tabela_vendas_carregamento.setRowCount(len(banco_compras))
        self.ui.tabela_vendas_carregamento.clear()

        colunas = ['Item', 'Cód', 'Produto', 'Valor De Custo', 'Qtde', 'Total']
        self.ui.tabela_vendas_carregamento.setHorizontalHeaderLabels(colunas)

        for compra in banco_compras:
            valor_de_custo = ''.join(re.findall(r'\d', compra[2]))
            valor_de_custo = lang.toString(int(valor_de_custo) * 0.01, 'f', 2)
            total = compra[4]

            self.ui.tabela_vendas_carregamento.setItem(row, 0, QTableWidgetItem(f'{compra[5]}'))
            self.ui.tabela_vendas_carregamento.setItem(row, 1, QTableWidgetItem(compra[0]))
            self.ui.tabela_vendas_carregamento.setItem(row, 2, QTableWidgetItem(compra[1]))
            self.ui.tabela_vendas_carregamento.setItem(row, 3, QTableWidgetItem('R$ ' + valor_de_custo))
            self.ui.tabela_vendas_carregamento.setItem(row, 4, QTableWidgetItem(compra[3]))
            self.ui.tabela_vendas_carregamento.setItem(row, 5, QTableWidgetItem('R$ ' + total))

            row += 1

        # Definir uma altura fixa para todas as linhas da tabela
        row_count = self.ui.tabela_vendas_carregamento.rowCount()
        for row in range(row_count):
            self.ui.tabela_vendas_carregamento.setRowHeight(row, 50)  # Define a altura de cada linha para 50 pixels

        # Ajustar o tamanho da fonte para os itens da tabela
        font = QtGui.QFont()
        font.setPointSize(13)  # Ajuste o tamanho da fonte conforme necessário
        self.ui.tabela_vendas_carregamento.setFont(font)

        # Habilitar a quebra de linha automática para todas as células da tabela
        self.ui.tabela_vendas_carregamento.resizeRowsToContents()

        # Ajustar o tamanho da fonte para os cabeçalhos das colunas da tabela
        header_font = QtGui.QFont()
        header_font.setPointSize(13)  # Ajuste o tamanho da fonte dos cabeçalhos conforme necessário
        self.ui.tabela_vendas_carregamento.horizontalHeader().setFont(header_font)

        # Ajustar a altura do cabeçalho da coluna da tabela
        self.ui.tabela_vendas_carregamento.horizontalHeader().setFixedHeight(
            40)  # Ajuste a altura do cabeçalho conforme necessário


if __name__ == '__main__':
    # Carregando Planilha
    wb = load_workbook('baseexcel.xlsx')

    # Variáveis Globais
    futuroTexto = ''

    # Clique dos botões de ver senha
    click_cadastro_colaboradores = 0
    click_cadastro_sangria = 0
    click_alterar_colaboradores = 0

    # Id para alterar os itens das tabelas
    id_tabela_alterar = None
    id_alterar_Clientes = None
    id_alterar_fornecedores = None
    id_alterar_produtos = None

    # Lista com os itens para fazer as previções das barras
    search_fornecedores = list()
    search_produtos = list()
    search_colaboradores = list()
    search_clientes = list()
    search_monitoramento = list()
    search_vendas = list()
    search_monitoramento_compras = list()
    search_compras = list()

    # Lista das Vendas
    vendas = list()

    # Lista das compras
    compras = list()

    # Conversção para moeda real (R$)
    loc = QLocale.system().name()
    lang = QLocale(loc)

    # Var para pegar o nome de quem logou
    UserLogado = None

    # Estilo de Erro e Padrão
    StyleError = '''
               background-color: rgba(0, 0 , 0, 0);
               border: 2px solid rgba(0,0,0,0);
               border-bottom-color: rgb(255, 17, 49);;
               color: rgb(0,0,0);
               padding-bottom: 8px;
               border-radius: 0px;
               font: 15pt "Montserrat";'''

    StyleNormal = '''
                   background-color: rgba(0, 0 , 0, 0);
                   border: 2px solid rgba(0,0,0,0);
                   border-bottom-color: rgb(12, 247, 28);;
                   color: rgb(0,0,0);
                   padding-bottom: 8px;
                   border-radius: 0px;
                   font: 15pt "Montserrat";'''

    # Configurando Aplicação
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('LogoIco.ico'))
    window = FrmLogin()
    window.show()
    sys.exit(app.exec_())

    # Fechando a conexão com o banco de dados quando o programa terminar
    banco.commit()
    cursor.close()
    banco.close()